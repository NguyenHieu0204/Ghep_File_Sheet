import os
import sys
import json
import urllib.error
import urllib.request
import tempfile
import threading
import tkinter as tk
from copy import copy
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
import win32com.client
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter as openpyxl_get_col_letter

CURRENT_VERSION = "1.4"
UPDATE_JSON_URL = "https://raw.githubusercontent.com/NguyenHieu0204/Ghep_File_Sheet/master/version.json"
UPDATE_JSON_URLS = [
    UPDATE_JSON_URL,
    "https://raw.githubusercontent.com/NguyenHieu0204/Ghep_File_Sheet/main/version.json",
]

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

XL_OPENXML_WORKBOOK = 51
XL_VALUES = -4163
XL_FORMATS = -4122
XL_FORMULAS = -4123
XL_BY_ROWS = 1
XL_BY_COLUMNS = 2
XL_PREVIOUS = 2


def parse_version(version):
    parts = []
    for part in str(version).split("."):
        number = ""
        for char in part:
            if char.isdigit():
                number += char
            else:
                break
        parts.append(int(number or 0))
    return tuple(parts)


def is_newer_version(latest_version, current_version):
    latest_parts = parse_version(latest_version)
    current_parts = parse_version(current_version)
    max_len = max(len(latest_parts), len(current_parts))
    latest_parts += (0,) * (max_len - len(latest_parts))
    current_parts += (0,) * (max_len - len(current_parts))
    return latest_parts > current_parts


def get_column_letter(n):
    """Convert zero-based column index to Excel-style letter."""
    result = ""
    while n >= 0:
        result = chr(65 + (n % 26)) + result
        n = (n // 26) - 1
    return result


def get_column_index(letter):
    """Convert Excel-style letter to zero-based column index."""
    n = 0
    for char in letter.upper():
        n = n * 26 + (ord(char) - 64)
    return n - 1


def normalize_xlsx_output_path(output_path):
    root, ext = os.path.splitext(output_path)
    return output_path if ext.lower() == ".xlsx" else root + ".xlsx"


def ensure_xlsx(file_path):
    """Convert .xls to .xlsx with Excel COM when needed."""
    if not str(file_path).lower().endswith(".xls"):
        return file_path

    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        abs_path = os.path.abspath(file_path)
        wb = excel.Workbooks.Open(abs_path)
        xlsx_path = os.path.join(tempfile.gettempdir(), os.path.basename(file_path) + "x")
        if os.path.exists(xlsx_path):
            os.remove(xlsx_path)
        wb.SaveAs(xlsx_path, FileFormat=XL_OPENXML_WORKBOOK)
        return xlsx_path
    except Exception as exc:
        raise RuntimeError(
            "Khong the tu dong chuyen doi file .xls sang .xlsx. "
            "Vui long mo file bang Excel va luu lai duoi dang .xlsx.\n"
            f"Chi tiet loi: {exc}"
        ) from exc
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def get_workbook_sheet_names(file_path):
    converted_path = ensure_xlsx(file_path)
    wb = load_workbook(converted_path, read_only=True, data_only=False)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def get_header_columns(file_path, sheet_name, header_row):
    converted_path = ensure_xlsx(file_path)
    wb = load_workbook(converted_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        values = []
        try:
            row_iter = ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
            row_values = next(row_iter)
            for index, cell_value in enumerate(row_values, start=1):
                title = cell_value if cell_value is not None else f"Column {index}"
                values.append(f"{get_column_letter(index - 1)} | {title}")
        except StopIteration:
            pass
        return values
    finally:
        wb.close()


def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        # Style indexes belong to a workbook, so register each distinct source
        # style once in the target workbook and safely reuse the registered style.
        target_wb = target_cell.parent.parent
        style_cache = getattr(target_wb, "_excel_tool_style_cache", None)
        if style_cache is None:
            style_cache = {}
            target_wb._excel_tool_style_cache = style_cache
        source_wb = source_cell.parent.parent
        source_token = getattr(source_wb, "_excel_tool_style_source_token", None)
        if source_token is None:
            source_token = object()
            source_wb._excel_tool_style_source_token = source_token
        cache_key = (source_token, source_cell.style_id)
        cached_style = style_cache.get(cache_key)
        if cached_style is None:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            cached_style = copy(target_cell._style)
            style_cache[cache_key] = cached_style
        else:
            target_cell._style = copy(cached_style)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def get_real_max_row(ws):
    """Return the last row that really has content, not just worksheet formatting."""
    real_max = 0
    for merged_range in ws.merged_cells.ranges:
        real_max = max(real_max, merged_range.max_row)

    if hasattr(ws, "_cells"):
        for (row_idx, _), cell in ws._cells.items():
            if cell.value is not None and str(cell.value).strip() != "":
                real_max = max(real_max, row_idx)
    else:
        real_max = ws.max_row
    return max(1, real_max)


def copy_sheet_data(source_ws, target_ws, start_row, include_header=True, header_row_count=1, fast_copy=False):
    row_offset = start_row - 1
    source_start_row = 1 if include_header else header_row_count + 1

    for col_letter, col_dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dimension.width
        target_ws.column_dimensions[col_letter].hidden = col_dimension.hidden

    for row_idx, row_dimension in source_ws.row_dimensions.items():
        if row_idx >= source_start_row:
            new_row_idx = row_idx + row_offset - (0 if include_header else header_row_count)
            target_ws.row_dimensions[new_row_idx].height = row_dimension.height
            target_ws.row_dimensions[new_row_idx].hidden = row_dimension.hidden

    real_max_row = get_real_max_row(source_ws)
    for row in source_ws.iter_rows(min_row=source_start_row, max_row=real_max_row):
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            target_row = cell.row + row_offset - (0 if include_header else header_row_count)
            new_cell = target_ws.cell(row=target_row, column=cell.column, value=cell.value)
            if not fast_copy:
                copy_cell_style(cell, new_cell)

    if not fast_copy:
        for merged_range in source_ws.merged_cells.ranges:
            if merged_range.min_row >= source_start_row:
                new_min_row = merged_range.min_row + row_offset - (0 if include_header else header_row_count)
                new_max_row = merged_range.max_row + row_offset - (0 if include_header else header_row_count)
                new_range = (
                    f"{openpyxl_get_col_letter(merged_range.min_col)}{new_min_row}:"
                    f"{openpyxl_get_col_letter(merged_range.max_col)}{new_max_row}"
                )
                try:
                    target_ws.merge_cells(new_range)
                except Exception:
                    pass

    return get_real_max_row(target_ws)


def get_excel_used_bounds(ws):
    last_row = 0
    last_col = 0

    try:
        last_row_cell = ws.Cells.Find(
            What="*",
            After=ws.Cells(1, 1),
            LookIn=XL_FORMULAS,
            SearchOrder=XL_BY_ROWS,
            SearchDirection=XL_PREVIOUS,
            MatchCase=False,
        )
        last_col_cell = ws.Cells.Find(
            What="*",
            After=ws.Cells(1, 1),
            LookIn=XL_FORMULAS,
            SearchOrder=XL_BY_COLUMNS,
            SearchDirection=XL_PREVIOUS,
            MatchCase=False,
        )
        if last_row_cell is not None:
            last_row = max(last_row, last_row_cell.Row)
        if last_col_cell is not None:
            last_col = max(last_col, last_col_cell.Column)
    except Exception:
        pass

    try:
        used = ws.UsedRange
        if used is not None:
            last_row = max(last_row, used.Row + used.Rows.Count - 1)
            last_col = max(last_col, used.Column + used.Columns.Count - 1)
    except Exception:
        pass

    return last_row, last_col


def prepare_destination_workbook(excel, sheet_title):
    wb = excel.Workbooks.Add()
    excel.DisplayAlerts = False
    while wb.Worksheets.Count > 1:
        wb.Worksheets(wb.Worksheets.Count).Delete()
    ws = wb.Worksheets(1)
    ws.Name = sheet_title
    return wb, ws


def copy_excel_row_properties(source_ws, target_ws, source_start_row, source_end_row, target_start_row):
    """Copy row height/visibility in bulk when uniform, with exact fallback when mixed."""
    if source_end_row < source_start_row:
        return
    row_count = source_end_row - source_start_row + 1
    source_rows = source_ws.Rows(f"{source_start_row}:{source_end_row}")
    target_rows = target_ws.Rows(f"{target_start_row}:{target_start_row + row_count - 1}")

    try:
        uniform_height = source_rows.RowHeight
    except Exception:
        uniform_height = None
    if uniform_height is not None:
        target_rows.RowHeight = uniform_height
    else:
        for offset in range(row_count):
            target_ws.Rows(target_start_row + offset).RowHeight = source_ws.Rows(source_start_row + offset).RowHeight

    try:
        uniform_hidden = source_rows.Hidden
    except Exception:
        uniform_hidden = None
    if uniform_hidden is not None:
        target_rows.Hidden = uniform_hidden
    else:
        for offset in range(row_count):
            target_ws.Rows(target_start_row + offset).Hidden = source_ws.Rows(source_start_row + offset).Hidden


def copy_excel_range(
    source_ws,
    target_ws,
    source_start_row,
    target_start_row,
    paste_values=False,
    used_bounds=None,
    fast_copy=False,
):
    last_row, last_col = used_bounds if used_bounds is not None else get_excel_used_bounds(source_ws)
    if last_row < source_start_row or last_col < 1:
        return target_start_row - 1

    source_range = source_ws.Range(source_ws.Cells(source_start_row, 1), source_ws.Cells(last_row, last_col))
    target_cell = target_ws.Cells(target_start_row, 1)

    if fast_copy:
        source_range.Copy()
        target_cell.PasteSpecial(Paste=XL_VALUES)
    elif paste_values:
        source_range.Copy()
        target_cell.PasteSpecial(Paste=XL_VALUES)
        target_cell.PasteSpecial(Paste=XL_FORMATS)
    else:
        source_range.Copy(Destination=target_cell)

    row_count = last_row - source_start_row + 1
    for col_idx in range(1, last_col + 1):
        target_ws.Columns(col_idx).ColumnWidth = source_ws.Columns(col_idx).ColumnWidth
        target_ws.Columns(col_idx).Hidden = source_ws.Columns(col_idx).Hidden
    copy_excel_row_properties(source_ws, target_ws, source_start_row, last_row, target_start_row)

    return target_start_row + row_count - 1


def copy_excel_fixed_range(
    source_ws,
    target_ws,
    source_start_row,
    source_end_row,
    target_start_row,
    paste_values=False,
    last_col=None,
    fast_copy=False,
):
    if source_end_row < source_start_row:
        return target_start_row - 1

    if last_col is None:
        _, last_col = get_excel_used_bounds(source_ws)
    if last_col < 1:
        return target_start_row - 1

    source_range = source_ws.Range(source_ws.Cells(source_start_row, 1), source_ws.Cells(source_end_row, last_col))
    target_cell = target_ws.Cells(target_start_row, 1)

    if fast_copy:
        source_range.Copy()
        target_cell.PasteSpecial(Paste=XL_VALUES)
    elif paste_values:
        source_range.Copy()
        target_cell.PasteSpecial(Paste=XL_VALUES)
        target_cell.PasteSpecial(Paste=XL_FORMATS)
    else:
        source_range.Copy(Destination=target_cell)

    row_count = source_end_row - source_start_row + 1
    for col_idx in range(1, last_col + 1):
        target_ws.Columns(col_idx).ColumnWidth = source_ws.Columns(col_idx).ColumnWidth
        target_ws.Columns(col_idx).Hidden = source_ws.Columns(col_idx).Hidden
    copy_excel_row_properties(source_ws, target_ws, source_start_row, source_end_row, target_start_row)

    return target_start_row + row_count - 1


def get_excel_last_display_row(source_ws, start_row=1):
    try:
        last_cell = source_ws.Cells.Find(
            What="*",
            After=source_ws.Cells(1, 1),
            LookIn=XL_VALUES,
            SearchOrder=XL_BY_ROWS,
            SearchDirection=XL_PREVIOUS,
            MatchCase=False,
        )
        if last_cell is not None and last_cell.Row >= start_row:
            return last_cell.Row
    except Exception:
        pass

    used = source_ws.UsedRange
    last_row = used.Row + used.Rows.Count - 1
    last_col = used.Column + used.Columns.Count - 1

    for row_idx in range(last_row, start_row - 1, -1):
        for col_idx in range(1, last_col + 1):
            text = source_ws.Cells(row_idx, col_idx).Text
            if text is not None and str(text).strip() != "":
                return row_idx
    return start_row - 1


def set_first_column_sequence_excel_com(target_ws, start_row, end_row, start_value):
    row_count = end_row - start_row + 1
    if row_count <= 0:
        return start_value

    values = tuple((value,) for value in range(start_value, start_value + row_count))
    target_ws.Range(target_ws.Cells(start_row, 1), target_ws.Cells(end_row, 1)).Value = values
    return start_value + row_count


def replace_summary_sheet(workbook, sheet_name):
    excel = workbook.Application
    excel.DisplayAlerts = False
    for idx in range(workbook.Worksheets.Count, 0, -1):
        if workbook.Worksheets(idx).Name == sheet_name:
            workbook.Worksheets(idx).Delete()
            break
    ws = workbook.Worksheets.Add(Before=workbook.Worksheets(1))
    ws.Name = sheet_name
    return ws


def merge_selected_sheets_to_summary_excel_com(
    input_path,
    output_path,
    selected_sheet_names,
    header_start_row,
    header_end_row,
    data_start_row,
    data_end_row=0,
    paste_values=False,
    renumber_first_column=True,
    summary_sheet_name="TongHop",
    fast_copy=False,
):
    if not selected_sheet_names:
        return 0

    excel = None
    wb = None
    try:
        converted_path = ensure_xlsx(input_path)
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(converted_path), ReadOnly=False)
        excel.ScreenUpdating = False
        excel.Calculation = -4135
        excel.EnableEvents = False
        summary_ws = replace_summary_sheet(wb, summary_sheet_name)

        first_source_ws = wb.Worksheets(selected_sheet_names[0])
        current_row = 1
        current_row = copy_excel_fixed_range(
            first_source_ws,
            summary_ws,
            header_start_row,
            header_end_row,
            current_row,
            paste_values,
            fast_copy=fast_copy,
        )
        current_row += 1

        sequence_number = 1
        for sheet_name in selected_sheet_names:
            source_ws = wb.Worksheets(sheet_name)
            source_bounds = get_excel_used_bounds(source_ws)
            source_end_row = data_end_row if data_end_row and data_end_row >= data_start_row else get_excel_last_display_row(
                source_ws, data_start_row
            )
            if source_end_row < data_start_row:
                continue

            target_start_row = current_row
            current_row = copy_excel_fixed_range(
                source_ws,
                summary_ws,
                data_start_row,
                source_end_row,
                target_start_row,
                paste_values,
                source_bounds[1],
                fast_copy,
            )

            if renumber_first_column:
                sequence_number = set_first_column_sequence_excel_com(
                    summary_ws, target_start_row, current_row, sequence_number
                )
            current_row += 1

        output_path = normalize_xlsx_output_path(output_path)
        if os.path.exists(output_path):
            os.remove(output_path)
        wb.SaveAs(os.path.abspath(output_path), FileFormat=XL_OPENXML_WORKBOOK)
        return len(selected_sheet_names)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        if excel is not None:
            try:
                excel.CutCopyMode = False
            except Exception:
                pass
            try: excel.ScreenUpdating = True
            except Exception: pass
            try: excel.Calculation = -4105
            except Exception: pass
            try: excel.EnableEvents = True
            except Exception: pass
            excel.Quit()


def merge_selected_sheets_excel_com(
    input_path,
    output_path,
    selected_sheet_names,
    keep_all_headers=True,
    paste_values=False,
    header_row_count=1,
    fast_copy=False,
):
    if not selected_sheet_names:
        return 0

    excel = None
    src_wb = None
    dst_wb = None
    try:
        converted_path = ensure_xlsx(input_path)
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        src_wb = excel.Workbooks.Open(os.path.abspath(converted_path), ReadOnly=True)
        excel.ScreenUpdating = False
        excel.Calculation = -4135
        excel.EnableEvents = False
        dst_wb, dst_ws = prepare_destination_workbook(excel, "Merged_Sheets")

        current_row = 1
        for index, sheet_name in enumerate(selected_sheet_names):
            source_ws = src_wb.Worksheets(sheet_name)
            source_bounds = get_excel_used_bounds(source_ws)
            include_header = keep_all_headers or index == 0
            source_start_row = 1 if include_header else header_row_count + 1
            current_row = copy_excel_range(
                source_ws, dst_ws, source_start_row, current_row, paste_values, source_bounds, fast_copy
            )
            current_row += 1

        output_path = normalize_xlsx_output_path(output_path)
        if os.path.exists(output_path):
            os.remove(output_path)
        dst_wb.SaveAs(os.path.abspath(output_path), FileFormat=XL_OPENXML_WORKBOOK)
        return len(selected_sheet_names)
    finally:
        if src_wb is not None:
            src_wb.Close(SaveChanges=False)
        if dst_wb is not None:
            dst_wb.Close(SaveChanges=False)
        if excel is not None:
            try:
                excel.CutCopyMode = False
            except Exception:
                pass
            try: excel.ScreenUpdating = True
            except Exception: pass
            try: excel.Calculation = -4105
            except Exception: pass
            try: excel.EnableEvents = True
            except Exception: pass
            excel.Quit()


def merge_multiple_files_excel_com(
    file_list,
    output_path,
    keep_all_headers=True,
    paste_values=False,
    header_start=1,
    header_end=1,
    fast_copy=False,
):
    if not file_list:
        return 0

    header_row_count = header_end - header_start + 1
    excel = None
    dst_wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        dst_wb, dst_ws = prepare_destination_workbook(excel, "Merged_Files")
        excel.ScreenUpdating = False
        excel.Calculation = -4135
        excel.EnableEvents = False

        current_row = 1
        source_index = 0
        for file_path in file_list:
            converted_path = ensure_xlsx(file_path)
            src_wb = excel.Workbooks.Open(os.path.abspath(converted_path), ReadOnly=True)
            try:
                for sheet_idx in range(1, src_wb.Worksheets.Count + 1):
                    source_ws = src_wb.Worksheets(sheet_idx)
                    source_bounds = get_excel_used_bounds(source_ws)
                    include_header = keep_all_headers or source_index == 0
                    if include_header:
                        # Copy header range (header_start..header_end)
                        current_row = copy_excel_fixed_range(
                            source_ws, dst_ws, header_start, header_end, current_row, paste_values, source_bounds[1], fast_copy
                        )
                        current_row += 1
                    # Copy data rows (after header_end)
                    current_row = copy_excel_range(
                        source_ws, dst_ws, header_end + 1, current_row, paste_values, source_bounds, fast_copy
                    )
                    current_row += 1
                    source_index += 1
            finally:
                src_wb.Close(SaveChanges=False)

        output_path = normalize_xlsx_output_path(output_path)
        if os.path.exists(output_path):
            os.remove(output_path)
        dst_wb.SaveAs(os.path.abspath(output_path), FileFormat=XL_OPENXML_WORKBOOK)
        return len(file_list)
    finally:
        if dst_wb is not None:
            dst_wb.Close(SaveChanges=False)
        if excel is not None:
            try:
                excel.CutCopyMode = False
            except Exception:
                pass
            try: excel.ScreenUpdating = True
            except Exception: pass
            try: excel.Calculation = -4105
            except Exception: pass
            try: excel.EnableEvents = True
            except Exception: pass
            excel.Quit()


def merge_selected_sheets_openpyxl(
    input_path,
    output_path,
    selected_sheet_names,
    keep_all_headers=True,
    header_row_count=1,
    fast_copy=False,
):
    if not selected_sheet_names:
        return 0

    converted_path = ensure_xlsx(input_path)
    src_wb = load_workbook(converted_path, data_only=False)
    dst_wb = Workbook()
    del dst_wb["Sheet"]
    dst_ws = dst_wb.create_sheet(title="Merged_Sheets")

    current_row = 1
    for index, sheet_name in enumerate(selected_sheet_names):
        source_ws = src_wb[sheet_name]
        include_header = keep_all_headers or index == 0
        current_row = copy_sheet_data(source_ws, dst_ws, current_row, include_header, header_row_count, fast_copy)
        current_row += 1

    src_wb.close()
    dst_wb.save(normalize_xlsx_output_path(output_path))
    return len(selected_sheet_names)


def merge_multiple_files_openpyxl(
    file_list,
    output_path,
    keep_all_headers=True,
    header_start=1,
    header_end=1,
    fast_copy=False,
):
    if not file_list:
        return 0

    header_row_count = header_end - header_start + 1
    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = "Merged_Files"
    current_row = 1
    source_index = 0

    for file_path in file_list:
        converted_path = ensure_xlsx(file_path)
        src_wb = load_workbook(converted_path, data_only=False)
        for sheet_name in src_wb.sheetnames:
            source_ws = src_wb[sheet_name]
            include_header = keep_all_headers or source_index == 0
            if include_header:
                # Copy header range (header_start..header_end)
                current_row = copy_row_range_openpyxl(
                    source_ws, dst_ws, header_start, header_end, current_row, fast_copy
                ) + 1
            # Copy data rows (after header_end)
            real_last = get_last_nonempty_row_openpyxl(source_ws, header_end + 1)
            if real_last >= header_end + 1:
                current_row = copy_row_range_openpyxl(
                    source_ws, dst_ws, header_end + 1, real_last, current_row, fast_copy
                )
            current_row += 1
            source_index += 1
        src_wb.close()

    dst_wb.save(normalize_xlsx_output_path(output_path))
    return len(file_list)


def merge_selected_sheets(
    input_path,
    output_path,
    selected_sheet_names,
    keep_all_headers=True,
    paste_values=False,
    prefer_excel_com=True,
    header_row_count=1,
    fast_copy=False,
):
    if prefer_excel_com:
        try:
            return merge_selected_sheets_excel_com(
                input_path, output_path, selected_sheet_names, keep_all_headers, paste_values, header_row_count, fast_copy
            )
        except Exception as exc:
            if paste_values:
                raise
            print(f"Excel COM failed, falling back to openpyxl: {exc}")
    return merge_selected_sheets_openpyxl(
        input_path, output_path, selected_sheet_names, keep_all_headers, header_row_count, fast_copy
    )


def merge_multiple_files(
    file_list,
    output_path,
    keep_all_headers=True,
    paste_values=False,
    prefer_excel_com=True,
    header_start=1,
    header_end=1,
    fast_copy=False,
):
    if prefer_excel_com:
        try:
            return merge_multiple_files_excel_com(
                file_list, output_path, keep_all_headers, paste_values, header_start, header_end, fast_copy
            )
        except Exception as exc:
            if paste_values:
                raise
            print(f"Excel COM failed, falling back to openpyxl: {exc}")
    return merge_multiple_files_openpyxl(file_list, output_path, keep_all_headers, header_start, header_end, fast_copy)


def copy_header_and_format(source_ws, target_ws, header_start, header_end, fast_copy=False):
    for col_letter, col_dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dimension.width
        target_ws.column_dimensions[col_letter].hidden = col_dimension.hidden

    target_row_idx = 1
    for row_idx in range(header_start, header_end + 1):
        if row_idx in source_ws.row_dimensions:
            target_ws.row_dimensions[target_row_idx].height = source_ws.row_dimensions[row_idx].height
            target_ws.row_dimensions[target_row_idx].hidden = source_ws.row_dimensions[row_idx].hidden

        for cell in source_ws[row_idx]:
            new_cell = target_ws.cell(row=target_row_idx, column=cell.column, value=cell.value)
            if not fast_copy:
                copy_cell_style(cell, new_cell)
        target_row_idx += 1

    if not fast_copy:
        for merged_range in source_ws.merged_cells.ranges:
            if merged_range.min_row >= header_start and merged_range.max_row <= header_end:
                new_min_row = merged_range.min_row - header_start + 1
                new_max_row = merged_range.max_row - header_start + 1
                new_range = (
                    f"{openpyxl_get_col_letter(merged_range.min_col)}{new_min_row}:"
                    f"{openpyxl_get_col_letter(merged_range.max_col)}{new_max_row}"
                )
                try:
                    target_ws.merge_cells(new_range)
                except Exception:
                    pass


def copy_data_chunk(source_ws, target_ws, rows_chunk, header_count, fast_copy=False):
    current_target_row = header_count + 1
    row_mapping = {}

    for row in rows_chunk:
        source_row_idx = row[0].row
        row_mapping[source_row_idx] = current_target_row

        if source_row_idx in source_ws.row_dimensions:
            target_ws.row_dimensions[current_target_row].height = source_ws.row_dimensions[source_row_idx].height
            target_ws.row_dimensions[current_target_row].hidden = source_ws.row_dimensions[source_row_idx].hidden

        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            new_cell = target_ws.cell(row=current_target_row, column=cell.column, value=cell.value)
            if not fast_copy:
                copy_cell_style(cell, new_cell)
        current_target_row += 1

    if not fast_copy:
        for merged_range in source_ws.merged_cells.ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            if min_row in row_mapping and max_row in row_mapping:
                new_range = (
                    f"{openpyxl_get_col_letter(merged_range.min_col)}{row_mapping[min_row]}:"
                    f"{openpyxl_get_col_letter(merged_range.max_col)}{row_mapping[max_row]}"
                )
                try:
                    target_ws.merge_cells(new_range)
                except Exception:
                    pass


def get_last_nonempty_row_openpyxl(ws, start_row=1):
    # Loaded worksheets keep instantiated cells in _cells. Scanning this sparse
    # mapping avoids materializing every empty cell up to ws.max_row/max_column.
    if hasattr(ws, "_cells"):
        candidates = (
            row_idx for (row_idx, _), cell in ws._cells.items()
            if row_idx >= start_row and cell.value is not None and str(cell.value).strip() != ""
        )
        return max(candidates, default=start_row - 1)
    for row_idx in range(ws.max_row, start_row - 1, -1):
        for cell in ws[row_idx]:
            if cell.value is not None and str(cell.value).strip() != "":
                return row_idx
    return start_row - 1


def copy_row_range_openpyxl(source_ws, target_ws, source_start_row, source_end_row, target_start_row, fast_copy=False):
    if source_end_row < source_start_row:
        return target_start_row - 1

    row_mapping = {}
    current_target_row = target_start_row
    for source_row_idx in range(source_start_row, source_end_row + 1):
        row_mapping[source_row_idx] = current_target_row

        if source_row_idx in source_ws.row_dimensions:
            target_ws.row_dimensions[current_target_row].height = source_ws.row_dimensions[source_row_idx].height
            target_ws.row_dimensions[current_target_row].hidden = source_ws.row_dimensions[source_row_idx].hidden

        for cell in source_ws[source_row_idx]:
            if cell.value is None and not cell.has_style:
                continue
            new_cell = target_ws.cell(row=current_target_row, column=cell.column, value=cell.value)
            if not fast_copy:
                copy_cell_style(cell, new_cell)
        current_target_row += 1

    if not fast_copy:
        for merged_range in source_ws.merged_cells.ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            if min_row in row_mapping and max_row in row_mapping:
                new_range = (
                    f"{openpyxl_get_col_letter(merged_range.min_col)}{row_mapping[min_row]}:"
                    f"{openpyxl_get_col_letter(merged_range.max_col)}{row_mapping[max_row]}"
                )
                try:
                    target_ws.merge_cells(new_range)
                except Exception:
                    pass

    return current_target_row - 1


def merge_selected_sheets_to_summary_openpyxl(
    input_path,
    output_path,
    selected_sheet_names,
    header_start_row,
    header_end_row,
    data_start_row,
    data_end_row=0,
    renumber_first_column=True,
    summary_sheet_name="TongHop",
    fast_copy=False,
):
    if not selected_sheet_names:
        return 0

    converted_path = ensure_xlsx(input_path)
    wb = load_workbook(converted_path, data_only=False)

    if summary_sheet_name in wb.sheetnames:
        del wb[summary_sheet_name]
    summary_ws = wb.create_sheet(title=summary_sheet_name, index=0)

    first_source_ws = wb[selected_sheet_names[0]]
    for col_letter, col_dimension in first_source_ws.column_dimensions.items():
        summary_ws.column_dimensions[col_letter].width = col_dimension.width
        summary_ws.column_dimensions[col_letter].hidden = col_dimension.hidden

    current_row = copy_row_range_openpyxl(
        first_source_ws, summary_ws, header_start_row, header_end_row, 1, fast_copy
    ) + 1

    sequence_number = 1
    for sheet_name in selected_sheet_names:
        source_ws = wb[sheet_name]
        source_end_row = data_end_row if data_end_row and data_end_row >= data_start_row else get_last_nonempty_row_openpyxl(
            source_ws, data_start_row
        )
        if source_end_row < data_start_row:
            continue

        target_start_row = current_row
        current_row = copy_row_range_openpyxl(
            source_ws, summary_ws, data_start_row, source_end_row, target_start_row, fast_copy
        )
        if renumber_first_column:
            for row_idx in range(target_start_row, current_row + 1):
                summary_ws.cell(row=row_idx, column=1).value = sequence_number
                sequence_number += 1
        current_row += 1

    wb.save(normalize_xlsx_output_path(output_path))
    wb.close()
    return len(selected_sheet_names)


def merge_selected_sheets_to_summary(
    input_path,
    output_path,
    selected_sheet_names,
    header_start_row,
    header_end_row,
    data_start_row,
    data_end_row=0,
    paste_values=False,
    prefer_excel_com=True,
    renumber_first_column=True,
    summary_sheet_name="TongHop",
    fast_copy=False,
):
    if prefer_excel_com:
        try:
            return merge_selected_sheets_to_summary_excel_com(
                input_path,
                output_path,
                selected_sheet_names,
                header_start_row,
                header_end_row,
                data_start_row,
                data_end_row,
                paste_values,
                renumber_first_column,
                summary_sheet_name,
                fast_copy,
            )
        except Exception as exc:
            if paste_values:
                raise
            print(f"Excel COM failed, falling back to openpyxl: {exc}")

    return merge_selected_sheets_to_summary_openpyxl(
        input_path,
        output_path,
        selected_sheet_names,
        header_start_row,
        header_end_row,
        data_start_row,
        data_end_row,
        renumber_first_column,
        summary_sheet_name,
        fast_copy,
    )


def clean_sheet_name(name):
    sheet_name = str(name)[:30].strip()
    for char in ["/", "\\", "?", "*", ":", "[", "]"]:
        sheet_name = sheet_name.replace(char, "")
    return sheet_name or "Sheet"


def split_excel_openpyxl(
    input_path,
    output_path_base,
    mode,
    output_mode,
    header_start,
    header_end,
    sheet_name=None,
    column_index=None,
    row_count=None,
    fast_copy=False,
):
    converted_path = ensure_xlsx(input_path)
    wb = load_workbook(converted_path, data_only=False)
    if not wb.sheetnames:
        return 0

    source_ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    real_max_row = get_real_max_row(source_ws)
    header_count = header_end - header_start + 1
    chunks = {}

    if mode == "column":
        for row in source_ws.iter_rows(min_row=header_end + 1, max_row=real_max_row):
            value = row[column_index].value
            key = str(value) if value is not None else "Blank"
            chunks.setdefault(key, []).append(row)
    else:
        current_chunk = 1
        current_rows = []
        for row in source_ws.iter_rows(min_row=header_end + 1, max_row=real_max_row):
            current_rows.append(row)
            if len(current_rows) >= row_count:
                chunks[f"Part_{current_chunk}"] = current_rows
                current_chunk += 1
                current_rows = []
        if current_rows:
            chunks[f"Part_{current_chunk}"] = current_rows

    if not chunks:
        return 0

    if output_mode == "sheets":
        for name, rows in chunks.items():
            safe_name = clean_sheet_name(name)
            base_name = safe_name
            counter = 1
            while safe_name in wb.sheetnames:
                safe_name = f"{base_name}_{counter}"
                counter += 1
            new_ws = wb.create_sheet(title=safe_name)
            copy_header_and_format(source_ws, new_ws, header_start, header_end, fast_copy)
            copy_data_chunk(source_ws, new_ws, rows, header_count, fast_copy)
        wb.save(normalize_xlsx_output_path(output_path_base))
        return len(chunks)

    base_dir = os.path.dirname(output_path_base)
    base_name_file, ext = os.path.splitext(os.path.basename(output_path_base))
    ext = ext or ".xlsx"
    for name, rows in chunks.items():
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = clean_sheet_name(name)
        copy_header_and_format(source_ws, new_ws, header_start, header_end, fast_copy)
        copy_data_chunk(source_ws, new_ws, rows, header_count, fast_copy)
        out_file = os.path.join(base_dir, f"{base_name_file}_{clean_sheet_name(name)}{ext}")
        new_wb.save(out_file)
    return len(chunks)



def split_excel_com(
    input_path,
    output_path_base,
    mode,
    output_mode,
    header_start,
    header_end,
    sheet_name=None,
    column_index=None,
    row_count=None,
    fast_copy=False,
):
    excel = None
    src_wb = None
    try:
        converted_path = ensure_xlsx(input_path)
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        src_wb = excel.Workbooks.Open(os.path.abspath(converted_path), ReadOnly=True)
        excel.ScreenUpdating = False
        excel.Calculation = -4135
        excel.EnableEvents = False
        try:
            source_ws = src_wb.Worksheets(sheet_name) if sheet_name else src_wb.Worksheets(1)
        except Exception:
            source_ws = src_wb.Worksheets(1)
            
        last_row, last_col = get_excel_used_bounds(source_ws)
        if last_row <= header_end:
            return 0
            
        base_dir = os.path.dirname(output_path_base)
        base_name_file, ext = os.path.splitext(os.path.basename(output_path_base))
        ext = ext or ".xlsx"
        
        num_parts = 0
        if mode == "column":
            col_letter = get_column_letter(column_index)
            val_range = source_ws.Range(f"{col_letter}{header_end + 1}:{col_letter}{last_row}").Value
            unique_vals = []
            seen = set()
            if val_range:
                for row_val in val_range:
                    val = row_val[0]
                    key = str(val) if val is not None else "Blank"
                    if key not in seen:
                        seen.add(key)
                        unique_vals.append(key)
            else:
                unique_vals.append("Blank")
            
            if not unique_vals:
                return 0
                
            if output_mode == "sheets":
                dst_wb = excel.Workbooks.Add()
                while dst_wb.Worksheets.Count > 1:
                    dst_wb.Worksheets(2).Delete()
            
            for i, key in enumerate(unique_vals):
                sheet_title = clean_sheet_name(key)
                if output_mode == "files":
                    dst_wb = excel.Workbooks.Add()
                    while dst_wb.Worksheets.Count > 1:
                        dst_wb.Worksheets(2).Delete()
                
                dst_ws = dst_wb.Worksheets(dst_wb.Worksheets.Count)
                dst_ws.Name = sheet_title
                
                copy_excel_fixed_range(
                    source_ws, dst_ws, header_start, header_end, 1, last_col=last_col, fast_copy=fast_copy
                )
                
                filter_range = source_ws.Range(source_ws.Cells(header_end, 1), source_ws.Cells(last_row, last_col))
                filter_field = column_index + 1
                filter_range.AutoFilter(Field=filter_field, Criteria1=key if key != "Blank" else "=")
                
                try:
                    data_range = source_ws.Range(source_ws.Cells(header_end + 1, 1), source_ws.Cells(last_row, last_col))
                    visible_range = data_range.SpecialCells(12)
                    target_cell = dst_ws.Cells(header_end - header_start + 2, 1)
                    if fast_copy:
                        visible_range.Copy()
                        target_cell.PasteSpecial(Paste=XL_VALUES)
                    else:
                        visible_range.Copy(Destination=target_cell)
                    
                    for col_idx in range(1, last_col + 1):
                        dst_ws.Columns(col_idx).ColumnWidth = source_ws.Columns(col_idx).ColumnWidth
                        dst_ws.Columns(col_idx).Hidden = source_ws.Columns(col_idx).Hidden
                except Exception:
                    pass
                    
                source_ws.AutoFilterMode = False
                
                if output_mode == "files":
                    out_file = os.path.join(base_dir, f"{base_name_file}_{sheet_title}{ext}")
                    if os.path.exists(out_file):
                        os.remove(out_file)
                    dst_wb.SaveAs(os.path.abspath(out_file), FileFormat=XL_OPENXML_WORKBOOK)
                    dst_wb.Close(SaveChanges=False)
                elif output_mode == "sheets" and i < len(unique_vals) - 1:
                    dst_wb.Worksheets.Add(After=dst_ws)
                    
            if output_mode == "sheets":
                dst_wb.SaveAs(os.path.abspath(output_path_base), FileFormat=XL_OPENXML_WORKBOOK)
                dst_wb.Close(SaveChanges=False)
                
            num_parts = len(unique_vals)
            
        else:
            current_start = header_end + 1
            current_chunk = 1
            
            if output_mode == "sheets":
                dst_wb = excel.Workbooks.Add()
                while dst_wb.Worksheets.Count > 1:
                    dst_wb.Worksheets(2).Delete()
                    
            while current_start <= last_row:
                current_end = min(current_start + row_count - 1, last_row)
                sheet_title = clean_sheet_name(f"Part_{current_chunk}")
                
                if output_mode == "files":
                    dst_wb = excel.Workbooks.Add()
                    while dst_wb.Worksheets.Count > 1:
                        dst_wb.Worksheets(2).Delete()
                        
                dst_ws = dst_wb.Worksheets(dst_wb.Worksheets.Count)
                dst_ws.Name = sheet_title
                
                copy_excel_fixed_range(
                    source_ws, dst_ws, header_start, header_end, 1, last_col=last_col, fast_copy=fast_copy
                )
                copy_excel_fixed_range(
                    source_ws,
                    dst_ws,
                    current_start,
                    current_end,
                    header_end - header_start + 2,
                    last_col=last_col,
                    fast_copy=fast_copy,
                )
                
                if output_mode == "files":
                    out_file = os.path.join(base_dir, f"{base_name_file}_{sheet_title}{ext}")
                    if os.path.exists(out_file):
                        os.remove(out_file)
                    dst_wb.SaveAs(os.path.abspath(out_file), FileFormat=XL_OPENXML_WORKBOOK)
                    dst_wb.Close(SaveChanges=False)
                elif output_mode == "sheets" and current_end < last_row:
                    dst_wb.Worksheets.Add(After=dst_ws)
                        
                current_chunk += 1
                current_start = current_end + 1
                num_parts += 1
                
            if output_mode == "sheets":
                dst_wb.SaveAs(os.path.abspath(output_path_base), FileFormat=XL_OPENXML_WORKBOOK)
                dst_wb.Close(SaveChanges=False)
                
        return num_parts
        
    finally:
        if src_wb is not None:
            src_wb.Close(SaveChanges=False)
        if excel is not None:
            try: excel.ScreenUpdating = True
            except Exception: pass
            try: excel.Calculation = -4105
            except Exception: pass
            try: excel.EnableEvents = True
            except Exception: pass
            try: excel.CutCopyMode = False
            except Exception: pass
            excel.Quit()


def split_excel(
    input_path,
    output_path_base,
    mode,
    output_mode,
    header_start,
    header_end,
    sheet_name=None,
    column_index=None,
    row_count=None,
    prefer_excel_com=True,
    fast_copy=False,
):
    if prefer_excel_com:
        try:
            return split_excel_com(
                input_path,
                output_path_base,
                mode,
                output_mode,
                header_start,
                header_end,
                sheet_name,
                column_index,
                row_count,
                fast_copy,
            )
        except Exception as exc:
            print(f"Excel COM failed, falling back to openpyxl: {exc}")
            
    return split_excel_openpyxl(
        input_path,
        output_path_base,
        mode,
        output_mode,
        header_start,
        header_end,
        sheet_name,
        column_index,
        row_count,
        fast_copy,
    )

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def normalize_reconcile_value(value, ignore_case=True):
    """Normalize Excel values so equivalent text/numbers can be compared safely."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    elif hasattr(value, "isoformat"):
        text = value.isoformat()
    else:
        text = str(value)
    text = " ".join(text.strip().split())
    return text.casefold() if ignore_case else text


def format_reconcile_result_message(
    source_columns, destination_columns, checked, matched, unique_source_keys, saved_path
):
    """Describe reconciliation counts without confusing columns with composite keys."""
    source_label = " + ".join(openpyxl_get_col_letter(index) for index in source_columns)
    destination_label = " + ".join(openpyxl_get_col_letter(index) for index in destination_columns)
    column_count = len(source_columns)
    return (
        f"Khóa đối soát: nguồn {source_label} ↔ đích {destination_label} "
        f"({column_count} cột mỗi bên).\n"
        f"Đã kiểm tra {checked} dòng đích; có {matched} dòng khớp.\n"
        f"Danh sách nguồn có {unique_source_keys} khóa ghép duy nhất.\n"
        f"Đã lưu: {saved_path}"
    )


def write_reconcile_batch_report(output_directory, report_rows):
    """Write one audit row per file pair, including success, error, or skipped status."""
    os.makedirs(output_directory, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(output_directory, f"BaoCao_DoiSoat_{timestamp}.xlsx")
    suffix = 2
    while os.path.exists(report_path):
        report_path = os.path.join(output_directory, f"BaoCao_DoiSoat_{timestamp}_{suffix}.xlsx")
        suffix += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "BaoCao_DoiSoat"
    headers = [
        "STT", "File nguồn", "Sheet nguồn", "File đích", "Sheet đích", "Trạng thái",
        "Dòng đã kiểm tra", "Dòng khớp", "Khóa nguồn duy nhất", "File kết quả", "Chi tiết lỗi",
    ]
    ws.append(headers)
    for row in report_rows:
        ws.append([
            row.get("index"), row.get("source"), row.get("source_sheet", ""),
            row.get("destination"), row.get("destination_sheet", ""), row.get("status"),
            row.get("checked", 0), row.get("matched", 0), row.get("source_keys", 0),
            row.get("output", ""), row.get("error", ""),
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = [8, 34, 24, 34, 24, 20, 18, 14, 22, 42, 60]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl_get_col_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(report_path)
    wb.close()
    return report_path


def reconcile_excel_files(
    source_path,
    destination_path,
    output_path,
    source_sheet,
    destination_sheet,
    source_columns,
    destination_columns,
    source_header_row=1,
    destination_header_row=1,
    note_header="Ghi chú đối soát",
    matched_note="Có",
    ignore_case=True,
    source_fill_column=None,
    fill_header="Dữ liệu đối soát",
    progress_callback=None,
):
    """Add a note column to a copy of the destination workbook for matched composite keys."""
    if not source_columns or len(source_columns) != len(destination_columns):
        raise ValueError("Số cột khóa của danh sách nguồn và đích phải bằng nhau.")
    if source_header_row < 1 or destination_header_row < 1:
        raise ValueError("Dòng tiêu đề phải lớn hơn hoặc bằng 1.")
    output_abs = os.path.normcase(os.path.abspath(normalize_xlsx_output_path(output_path)))
    source_abs = os.path.normcase(os.path.abspath(source_path))
    destination_abs = os.path.normcase(os.path.abspath(destination_path))
    if output_abs in (source_abs, destination_abs):
        raise ValueError("File kết quả phải khác file nguồn và file đích để bảo vệ dữ liệu gốc.")

    def report_progress(progress, status):
        if progress_callback:
            progress_callback(max(0.0, min(1.0, progress)), status)

    report_progress(0.0, "Đang mở và kiểm tra hai file Excel...")
    source_converted = ensure_xlsx(source_path)
    destination_converted = ensure_xlsx(destination_path)
    # Do not use read_only here: some Excel generators leave a stale worksheet
    # dimension (for example A1:V9 although data continues to row 1629). In that
    # case openpyxl read-only iteration silently stops early and misses keys.
    source_wb = load_workbook(source_converted, data_only=True)
    # Normal mode intentionally ignores stale worksheet dimension metadata that
    # some third-party exports write (for example A1:V9 while data reaches row 3006).
    destination_values_wb = load_workbook(destination_converted, data_only=True)
    destination_wb = load_workbook(destination_converted, data_only=False)
    try:
        if source_sheet not in source_wb.sheetnames:
            raise ValueError(f"Không tìm thấy sheet nguồn: {source_sheet}")
        if destination_sheet not in destination_wb.sheetnames or destination_sheet not in destination_values_wb.sheetnames:
            raise ValueError(f"Không tìm thấy sheet đích: {destination_sheet}")

        source_ws = source_wb[source_sheet]
        destination_values_ws = destination_values_wb[destination_sheet]
        destination_ws = destination_wb[destination_sheet]
        source_keys = set()
        source_fill_values = {}
        source_total = max(0, source_ws.max_row - source_header_row)
        source_step = max(1, source_total // 200)
        for source_index, row in enumerate(
            source_ws.iter_rows(min_row=source_header_row + 1, values_only=True), start=1
        ):
            key = tuple(
                normalize_reconcile_value(row[index - 1] if index <= len(row) else None, ignore_case)
                for index in source_columns
            )
            if any(key):
                source_keys.add(key)
                if source_fill_column:
                    fill_value = row[source_fill_column - 1] if source_fill_column <= len(row) else None
                    # Keep the first value, but allow a later non-empty value to replace an empty one.
                    if key not in source_fill_values or source_fill_values[key] in (None, ""):
                        source_fill_values[key] = fill_value
            if source_index == source_total or source_index % source_step == 0:
                ratio = source_index / max(1, source_total)
                report_progress(
                    0.25 * ratio,
                    f"Đang tạo khóa nguồn: {source_index:,}/{source_total:,} dòng",
                )

        note_column = destination_ws.max_column + 1
        header_cell = destination_ws.cell(destination_header_row, note_column, note_header.strip() or "Ghi chú đối soát")
        if note_column > 1:
            copy_cell_style(destination_ws.cell(destination_header_row, note_column - 1), header_cell)
            previous_width = destination_ws.column_dimensions[openpyxl_get_col_letter(note_column - 1)].width
            destination_ws.column_dimensions[openpyxl_get_col_letter(note_column)].width = max(previous_width or 10, 18)

        fill_column = None
        if source_fill_column:
            fill_column = note_column + 1
            fill_cell = destination_ws.cell(
                destination_header_row, fill_column, fill_header.strip() or "Dữ liệu đối soát"
            )
            copy_cell_style(header_cell, fill_cell)
            destination_ws.column_dimensions[openpyxl_get_col_letter(fill_column)].width = 20

        matched_count = 0
        checked_count = 0
        destination_total = max(0, destination_ws.max_row - destination_header_row)
        destination_step = max(1, destination_total // 300)
        for row_index in range(destination_header_row + 1, destination_ws.max_row + 1):
            key = tuple(
                normalize_reconcile_value(destination_values_ws.cell(row_index, index).value, ignore_case)
                for index in destination_columns
            )
            if not any(key):
                continue
            checked_count += 1
            if key in source_keys:
                destination_ws.cell(row_index, note_column, matched_note)
                if fill_column:
                    destination_ws.cell(row_index, fill_column, source_fill_values.get(key))
                matched_count += 1

            processed = row_index - destination_header_row
            if processed == destination_total or processed % destination_step == 0:
                ratio = processed / max(1, destination_total)
                report_progress(
                    0.25 + 0.65 * ratio,
                    f"Đã đối soát: {processed:,}/{destination_total:,} dòng • Khớp: {matched_count:,}",
                )

        output_path = normalize_xlsx_output_path(output_path)
        report_progress(0.92, "Đang lưu file kết quả, vui lòng chờ...")
        destination_wb.save(output_path)
        report_progress(1.0, "Hoàn tất đối soát và lưu file kết quả")
        return matched_count, checked_count, len(source_keys), output_path
    finally:
        source_wb.close()
        destination_values_wb.close()
        destination_wb.close()


class ExcelMergerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(f"Excel Tool - Ghep Sheet/File v{CURRENT_VERSION}")
        self.geometry("1100x700")
        try:
            self.iconbitmap(resource_path("app_icon.ico"))
        except:
            pass

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self.navigation_frame = ctk.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(6, weight=1)

        self.label_header = ctk.CTkLabel(
            self.navigation_frame,
            text="  ⚡ EXCEL TOOLBOX",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=("black", "white"),
        )
        self.label_header.grid(row=0, column=0, padx=20, pady=20)

        self.home_button = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="📋  Ghép Sheet (1 File)",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            command=lambda: self.select_frame_by_name("merge_sheet"),
        )
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.frame_2_button = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="📁  Ghép Nhiều File",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            command=lambda: self.select_frame_by_name("merge_files"),
        )
        self.frame_2_button.grid(row=2, column=0, sticky="ew")

        self.frame_3_button = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="✂  Tách File Excel",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            command=lambda: self.select_frame_by_name("split_file"),
        )
        self.frame_3_button.grid(row=3, column=0, sticky="ew")

        self.reconcile_button = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="🔎  Đối Soát Dữ Liệu",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            command=lambda: self.select_frame_by_name("reconcile"),
        )
        self.reconcile_button.grid(row=4, column=0, sticky="ew")

        self.frame_4_button = ctk.CTkButton(
            self.navigation_frame, corner_radius=0, height=40, border_spacing=10,
            text="📖  Hướng Dẫn Sử Dụng", fg_color="transparent",
            text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), anchor="w",
            command=lambda: self.select_frame_by_name("guide"),
        )
        self.frame_4_button.grid(row=5, column=0, sticky="ew")

        version_frame = ctk.CTkFrame(self.navigation_frame, fg_color="transparent")
        version_frame.grid(row=7, column=0, padx=12, pady=14, sticky="ew")
        self.version_label = ctk.CTkLabel(
            version_frame, text=f"Phiên bản {CURRENT_VERSION}",
            font=ctk.CTkFont(size=12, weight="bold"), text_color=("gray35", "gray70"),
        )
        self.version_label.pack(pady=(0, 5))
        self.update_status_label = ctk.CTkLabel(
            version_frame, text="", font=ctk.CTkFont(size=11), text_color=("gray40", "gray65")
        )
        self.update_status_label.pack()
        self.check_update_button = ctk.CTkButton(
            version_frame, text="Kiểm tra cập nhật", height=28, width=150,
            command=self.check_for_updates_manually,
        )
        self.check_update_button.pack(pady=(5, 0))

        self.home_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.second_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.third_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.reconcile_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.guide_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")

        self.setup_home_frame()
        self.setup_second_frame()
        self.setup_third_frame()
        self.setup_reconcile_frame()
        self.setup_guide_frame()
        self.select_frame_by_name("merge_sheet")
        threading.Thread(target=self.check_for_updates, daemon=True).start()

    def check_for_updates_manually(self):
        self.check_update_button.configure(state="disabled")
        self.update_status_label.configure(text="Đang kiểm tra...", text_color="#f39c12")
        threading.Thread(target=self.check_for_updates, args=(True,), daemon=True).start()

    def check_for_updates(self, show_result=False):
        connected = False
        for update_url in UPDATE_JSON_URLS:
            try:
                req = urllib.request.Request(update_url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=5) as response:
                    data = json.loads(response.read().decode("utf-8"))
                    connected = True
                    latest_version = data.get("version", CURRENT_VERSION)
                    download_url = data.get("download_url", "")

                    if is_newer_version(latest_version, CURRENT_VERSION):
                        self.after(0, lambda v=latest_version: self.update_status_label.configure(
                            text=f"Có bản mới {v}", text_color="#e67e22"
                        ))
                        if download_url:
                            self.after(0, lambda v=latest_version, u=download_url: self.prompt_update(v, u))
                        elif show_result:
                            self.after(0, lambda v=latest_version: messagebox.showinfo(
                                "Có phiên bản mới",
                                f"Phiên bản hiện tại: {CURRENT_VERSION}\nPhiên bản mới: {v}\n\n"
                                "Máy chủ chưa cung cấp đường dẫn tải xuống.",
                            ))
                    else:
                        self.after(0, lambda: self.update_status_label.configure(
                            text="Đang dùng bản mới nhất", text_color="#2ecc71"
                        ))
                        if show_result:
                            self.after(0, lambda: messagebox.showinfo(
                                "Kiểm tra cập nhật",
                                f"Bạn đang sử dụng phiên bản mới nhất ({CURRENT_VERSION}).",
                            ))
                    self.after(0, lambda: self.check_update_button.configure(state="normal"))
                    return
            except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, json.JSONDecodeError):
                continue
            except Exception:
                continue
        self.after(0, lambda: self.update_status_label.configure(
            text="Không thể kiểm tra cập nhật", text_color="#e74c3c"
        ))
        self.after(0, lambda: self.check_update_button.configure(state="normal"))
        if show_result:
            self.after(0, lambda: messagebox.showwarning(
                "Không thể kiểm tra cập nhật",
                "Không thể kết nối máy chủ cập nhật. Vui lòng kiểm tra kết nối Internet và thử lại.",
            ))

    def prompt_update(self, latest_version, download_url):
        result = messagebox.askyesno(
            "Có phiên bản mới",
            f"Phiên bản hiện tại: {CURRENT_VERSION}\n"
            f"Phiên bản mới: {latest_version}\n\n"
            "Bạn có muốn tải xuống và cài đặt bản mới ngay không?"
        )
        if result:
            self.update_status_label.configure(text="Đang tải bản cập nhật...", text_color="#3498db")
            threading.Thread(target=self.download_and_install_update, args=(download_url,), daemon=True).start()

    def download_and_install_update(self, download_url):
        try:
            exe_path = os.path.join(tempfile.gettempdir(), "Excel_Toolbox_Update.exe")
            urllib.request.urlretrieve(download_url, exe_path)
            os.startfile(exe_path)
            self.after(1000, self.destroy)
        except Exception as e:
            self.after(0, lambda: self.update_status_label.configure(
                text="Tải cập nhật thất bại", text_color="#e74c3c"
            ))
            self.after(0, lambda: messagebox.showerror("Lỗi cập nhật", f"Không thể tải bản cập nhật:\n{e}"))

    def setup_merge_options(self, parent, row, column=0, columnspan=1):
        options = ctk.CTkFrame(parent, fg_color="transparent")
        options.grid(row=row, column=column, columnspan=columnspan, padx=20, pady=5, sticky="w")
        keep_headers_var = ctk.BooleanVar(value=True)
        paste_values_var = ctk.BooleanVar(value=False)
        prefer_com_var = ctk.BooleanVar(value=True)
        fast_copy_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(options, text="Giữ tiêu đề của mỗi sheet/file", variable=keep_headers_var).grid(row=0, column=0, padx=8)
        ctk.CTkCheckBox(options, text="Dán giá trị thay vì công thức", variable=paste_values_var).grid(row=0, column=1, padx=8)
        ctk.CTkCheckBox(options, text="Ưu tiên Excel COM (Giữ format tốt nhất)", variable=prefer_com_var).grid(row=0, column=2, padx=8)
        ctk.CTkCheckBox(options, text="Copy nhanh (khong giu full format)", variable=fast_copy_var).grid(row=1, column=0, columnspan=3, padx=8, pady=(5, 0), sticky="w")
        return keep_headers_var, paste_values_var, prefer_com_var, fast_copy_var

    def setup_summary_merge_options(self, parent, row, column=0, columnspan=1):
        options = ctk.CTkFrame(parent, fg_color="transparent")
        options.grid(row=row, column=column, columnspan=columnspan, padx=20, pady=5, sticky="w")
        paste_values_var = ctk.BooleanVar(value=False)
        prefer_com_var = ctk.BooleanVar(value=True)
        renumber_var = ctk.BooleanVar(value=True)
        fast_copy_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(options, text="Dán giá trị thay vì công thức", variable=paste_values_var).grid(row=0, column=0, padx=8)
        ctk.CTkCheckBox(options, text="Ưu tiên Excel COM (Giữ format tốt nhất)", variable=prefer_com_var).grid(row=0, column=1, padx=8)
        ctk.CTkCheckBox(options, text="Đánh lại STT cột A", variable=renumber_var).grid(row=0, column=2, padx=8)
        ctk.CTkCheckBox(options, text="Copy nhanh (khong giu full format)", variable=fast_copy_var).grid(row=1, column=0, columnspan=3, padx=8, pady=(5, 0), sticky="w")
        return paste_values_var, prefer_com_var, renumber_var, fast_copy_var

    def setup_home_frame(self):
        ctk.CTkLabel(
            self.home_frame,
            text="📋  Ghép Sheet (1 File) vào Tổng Hợp",
            font=ctk.CTkFont(size=20, weight="bold"),
        ).grid(row=0, column=0, padx=20, pady=20)
        self.input_file_path = ctk.StringVar()

        file_frame = ctk.CTkFrame(self.home_frame, fg_color="transparent")
        file_frame.grid(row=1, column=0, padx=20, pady=10)
        ctk.CTkEntry(file_frame, textvariable=self.input_file_path, width=400).grid(row=0, column=0, padx=5)
        ctk.CTkButton(file_frame, text="📂  Chọn File", command=self.browse_single_file).grid(row=0, column=1, padx=5)

        self.sheet_info_label = ctk.CTkLabel(self.home_frame, text="Chọn các sheet cần ghép:", text_color=("gray40", "gray60"))
        self.sheet_info_label.grid(row=2, column=0, pady=(10, 0))

        list_frame = ctk.CTkFrame(self.home_frame)
        list_frame.grid(row=3, column=0, padx=20, pady=5)
        self.home_sheet_listbox = tk.Listbox(list_frame, selectmode="multiple", width=60, height=8, bg="#333333", fg="white", selectbackground="#2ecc71")
        self.home_sheet_listbox.pack(side="left", fill="both", expand=True)
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=self.home_sheet_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.home_sheet_listbox.config(yscrollcommand=scrollbar.set)

        ctk.CTkButton(self.home_frame, text="☑  Chọn tất cả Sheet", command=self.select_all_home_sheets, width=150).grid(row=4, column=0, pady=5)

        range_frame = ctk.CTkFrame(self.home_frame, fg_color="transparent")
        range_frame.grid(row=5, column=0, padx=20, pady=5)
        ctk.CTkLabel(range_frame, text="Tiêu đề từ dòng:").grid(row=0, column=0, padx=2)
        self.merge_header_start_entry = ctk.CTkEntry(range_frame, width=50)
        self.merge_header_start_entry.insert(0, "1")
        self.merge_header_start_entry.grid(row=0, column=1, padx=2)
        ctk.CTkLabel(range_frame, text="đến:").grid(row=0, column=2, padx=2)
        self.merge_header_end_entry = ctk.CTkEntry(range_frame, width=50)
        self.merge_header_end_entry.insert(0, "7")
        self.merge_header_end_entry.grid(row=0, column=3, padx=2)
        ctk.CTkLabel(range_frame, text="Dữ liệu từ dòng:").grid(row=0, column=4, padx=(16, 2))
        self.merge_data_start_entry = ctk.CTkEntry(range_frame, width=50)
        self.merge_data_start_entry.insert(0, "8")
        self.merge_data_start_entry.grid(row=0, column=5, padx=2)
        ctk.CTkLabel(range_frame, text="đến:").grid(row=0, column=6, padx=2)
        self.merge_data_end_entry = ctk.CTkEntry(range_frame, width=70)
        self.merge_data_end_entry.insert(0, "0")
        self.merge_data_end_entry.grid(row=0, column=7, padx=2)
        ctk.CTkLabel(range_frame, text="0 = tự động tới dòng cuối có dữ liệu").grid(row=0, column=8, padx=8)

        (
            self.home_paste_values_var,
            self.home_prefer_com_var,
            self.home_renumber_var,
            self.home_fast_copy_var,
        ) = self.setup_summary_merge_options(self.home_frame, 6)
        self.home_merge_btn = ctk.CTkButton(self.home_frame, text="▶  Bắt đầu Ghép Sheet", command=self.merge_sheets_process, height=50, fg_color="#2ecc71", font=ctk.CTkFont(size=14, weight="bold"))
        self.home_merge_btn.grid(row=7, column=0, pady=20)
        self.home_progress_frame, self.home_progress_bar, self.home_status_label = self._create_progress_widgets(self.home_frame, row=8)

    def setup_second_frame(self):
        ctk.CTkLabel(
            self.second_frame,
            text="📁  Ghép nhiều file Excel",
            font=ctk.CTkFont(size=20, weight="bold"),
        ).grid(row=0, column=0, padx=20, pady=20)
        self.file_list = []
        self.file_listbox = tk.Listbox(self.second_frame, width=80, height=10, bg="#333333", fg="white")
        self.file_listbox.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        file_ops = ctk.CTkFrame(self.second_frame, fg_color="transparent")
        file_ops.grid(row=1, column=1, padx=20, pady=10, sticky="n")
        ctk.CTkButton(file_ops, text="➕  Thêm Files", command=self.browse_multiple_files).grid(row=0, column=0, pady=5)
        ctk.CTkButton(file_ops, text="🗑  Xóa", fg_color="#e74c3c", command=self.clear_file_list).grid(row=1, column=0, pady=5)

        # Header range inputs
        files_range_frame = ctk.CTkFrame(self.second_frame, fg_color="transparent")
        files_range_frame.grid(row=2, column=0, columnspan=2, padx=20, pady=5)
        ctk.CTkLabel(files_range_frame, text="Tiêu đề từ dòng:").grid(row=0, column=0, padx=2)
        self.files_header_start_entry = ctk.CTkEntry(files_range_frame, width=50)
        self.files_header_start_entry.insert(0, "1")
        self.files_header_start_entry.grid(row=0, column=1, padx=2)
        ctk.CTkLabel(files_range_frame, text="đến:").grid(row=0, column=2, padx=2)
        self.files_header_end_entry = ctk.CTkEntry(files_range_frame, width=50)
        self.files_header_end_entry.insert(0, "1")
        self.files_header_end_entry.grid(row=0, column=3, padx=2)
        ctk.CTkLabel(files_range_frame, text="(dòng dữ liệu bắt đầu từ dòng tiếp theo)", text_color=("gray40", "gray60")).grid(row=0, column=4, padx=8)

        (
            self.files_keep_headers_var,
            self.files_paste_values_var,
            self.files_prefer_com_var,
            self.files_fast_copy_var,
        ) = self.setup_merge_options(self.second_frame, 3, column=0, columnspan=2)
        self.files_merge_btn = ctk.CTkButton(self.second_frame, text="▶  Bắt đầu Ghép File", command=self.merge_files_process, height=50, fg_color="#3498db", font=ctk.CTkFont(size=14, weight="bold"))
        self.files_merge_btn.grid(row=4, column=0, columnspan=2, pady=20)
        self.files_progress_frame, self.files_progress_bar, self.files_status_label = self._create_progress_widgets(self.second_frame, row=5, columnspan=2)

    def setup_third_frame(self):
        ctk.CTkLabel(self.third_frame, text="✂  Tách File Excel", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, columnspan=2, padx=20, pady=20)
        self.split_input_path = ctk.StringVar()
        file_f = ctk.CTkFrame(self.third_frame, fg_color="transparent")
        file_f.grid(row=1, column=0, columnspan=2, padx=20, pady=10)
        ctk.CTkEntry(file_f, textvariable=self.split_input_path, width=400).grid(row=0, column=0, padx=5)
        ctk.CTkButton(file_f, text="📂  Chọn File", command=self.browse_split_file).grid(row=0, column=1, padx=5)

        sheet_f = ctk.CTkFrame(self.third_frame, fg_color="transparent")
        sheet_f.grid(row=2, column=0, columnspan=2, padx=20, pady=5)
        ctk.CTkLabel(sheet_f, text="Chọn Sheet cần tách:").grid(row=0, column=0, padx=5)
        self.split_sheet_var = ctk.StringVar(value="Active Sheet")
        self.split_sheet_menu = ctk.CTkOptionMenu(sheet_f, variable=self.split_sheet_var, values=["Active Sheet"], command=lambda _: self.reload_columns())
        self.split_sheet_menu.grid(row=0, column=1, padx=5)

        options_frame = ctk.CTkFrame(self.third_frame, fg_color="transparent")
        options_frame.grid(row=3, column=0, columnspan=2, pady=10)

        self.split_mode_var = ctk.StringVar(value="column")
        mode_frame = ctk.CTkFrame(options_frame)
        mode_frame.grid(row=0, column=0, padx=10, sticky="n")
        ctk.CTkLabel(mode_frame, text="Phương thức tách:", font=ctk.CTkFont(weight="bold")).pack(pady=5)
        ctk.CTkRadioButton(mode_frame, text="Theo giá trị cột", variable=self.split_mode_var, value="column", command=self.toggle_split_inputs).pack(anchor="w", padx=10, pady=5)
        ctk.CTkRadioButton(mode_frame, text="Theo số lượng dòng", variable=self.split_mode_var, value="row_count", command=self.toggle_split_inputs).pack(anchor="w", padx=10, pady=5)

        self.split_output_var = ctk.StringVar(value="sheets")
        output_frame = ctk.CTkFrame(options_frame)
        output_frame.grid(row=0, column=1, padx=10, sticky="n")
        ctk.CTkLabel(output_frame, text="Hình thức xuất:", font=ctk.CTkFont(weight="bold")).pack(pady=5)
        ctk.CTkRadioButton(output_frame, text="Nhiều Sheet trong 1 File", variable=self.split_output_var, value="sheets").pack(anchor="w", padx=10, pady=5)
        ctk.CTkRadioButton(output_frame, text="Nhiều File riêng biệt", variable=self.split_output_var, value="files").pack(anchor="w", padx=10, pady=5)

        self.inputs_frame = ctk.CTkFrame(self.third_frame, fg_color="transparent")
        self.inputs_frame.grid(row=4, column=0, columnspan=2, pady=5)
        ctk.CTkLabel(self.inputs_frame, text="Tiêu đề từ dòng:").grid(row=0, column=0, padx=2)
        self.header_start_entry = ctk.CTkEntry(self.inputs_frame, width=40)
        self.header_start_entry.insert(0, "1")
        self.header_start_entry.grid(row=0, column=1, padx=2)
        ctk.CTkLabel(self.inputs_frame, text="đến:").grid(row=0, column=2, padx=2)
        self.header_end_entry = ctk.CTkEntry(self.inputs_frame, width=40)
        self.header_end_entry.insert(0, "1")
        self.header_end_entry.grid(row=0, column=3, padx=2)

        self.split_header_menu = ctk.CTkOptionMenu(self.inputs_frame, values=["Chưa có dữ liệu"], width=180)
        self.split_header_menu.grid(row=0, column=4, padx=5)
        self.row_count_entry = ctk.CTkEntry(self.inputs_frame, width=150, placeholder_text="Số dòng mỗi phần...")
        self.row_count_entry.grid_remove()
        ctk.CTkButton(self.inputs_frame, text="🔄  Tải lại cột", width=90, command=self.reload_columns).grid(row=0, column=5, padx=5)
        self.split_fast_copy_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            self.third_frame,
            text="Copy nhanh (khong giu full format)",
            variable=self.split_fast_copy_var,
        ).grid(row=5, column=0, columnspan=2, pady=(5, 0))
        self.split_btn = ctk.CTkButton(self.third_frame, text="▶  Bắt đầu Tách File", command=self.split_process, height=50, fg_color="#9b59b6", font=ctk.CTkFont(size=14, weight="bold"))
        self.split_btn.grid(row=6, column=0, columnspan=2, pady=20)
        self.split_progress_frame, self.split_progress_bar, self.split_status_label = self._create_progress_widgets(self.third_frame, row=7, columnspan=2)

    def setup_reconcile_frame(self):
        self.reconcile_frame.grid_columnconfigure(0, weight=1)
        self.reconcile_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(
            self.reconcile_frame, text="🔎  Đối Soát Dữ Liệu Theo Nhiều Cột",
            font=ctk.CTkFont(size=20, weight="bold"),
        ).grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 8))

        self.reconcile_source_path = ctk.StringVar()
        self.reconcile_destination_path = ctk.StringVar()
        self.reconcile_source_files = []
        self.reconcile_destination_files = []
        self.reconcile_source_pair_sheets = []
        self.reconcile_destination_pair_sheets = []
        self.reconcile_source_sheet_options = []
        self.reconcile_destination_sheet_options = []
        self.reconcile_pairs_window = None
        self.reconcile_pairs_tree = None
        self.reconcile_source_sheet = ctk.StringVar(value="Chưa có dữ liệu")
        self.reconcile_destination_sheet = ctk.StringVar(value="Chưa có dữ liệu")

        for column, title, path_var, command in (
            (0, "NGUỒN – File dùng để tìm/tra cứu", self.reconcile_source_path, lambda: self.browse_reconcile_file("source")),
            (1, "ĐÍCH – File nhận cột ghi chú/kết quả", self.reconcile_destination_path, lambda: self.browse_reconcile_file("destination")),
        ):
            panel = ctk.CTkFrame(self.reconcile_frame)
            panel.grid(row=1, column=column, padx=10, pady=5, sticky="nsew")
            panel.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(panel, text=title, font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=2, pady=5)
            ctk.CTkEntry(panel, textvariable=path_var, width=310).grid(row=1, column=0, padx=5, pady=5, sticky="ew")
            file_buttons = ctk.CTkFrame(panel, fg_color="transparent")
            file_buttons.grid(row=1, column=1, padx=5)
            load_text = "📂 Tải nguồn" if column == 0 else "📂 Tải đích"
            ctk.CTkButton(file_buttons, text=load_text, width=92, command=command).grid(row=0, column=0, padx=2)
            side = "source" if column == 0 else "destination"
            ctk.CTkButton(
                file_buttons, text="✕ Xóa", width=62, fg_color="#c0392b", hover_color="#922b21",
                command=lambda selected_side=side: self.clear_reconcile_file(selected_side),
            ).grid(row=0, column=1, padx=2)

            sheet_var = self.reconcile_source_sheet if column == 0 else self.reconcile_destination_sheet
            sheet_menu = ctk.CTkOptionMenu(
                panel, variable=sheet_var, values=["Chưa có dữ liệu"], width=220,
                command=lambda value, side="source" if column == 0 else "destination":
                    self.on_reconcile_main_sheet_changed(side, value),
            )
            sheet_menu.grid(row=2, column=0, padx=5, pady=5, sticky="w")
            if column == 0:
                self.reconcile_source_sheet_menu = sheet_menu
            else:
                self.reconcile_destination_sheet_menu = sheet_menu

            header_range = ctk.CTkFrame(panel, fg_color="transparent")
            header_range.grid(row=2, column=1, padx=3, pady=5, sticky="e")
            ctk.CTkLabel(header_range, text="Tiêu đề:").grid(row=0, column=0, padx=2)
            header_start_entry = ctk.CTkEntry(header_range, width=38)
            header_start_entry.insert(0, "1")
            header_start_entry.grid(row=0, column=1, padx=2)
            ctk.CTkLabel(header_range, text="đến").grid(row=0, column=2, padx=2)
            header_entry = ctk.CTkEntry(header_range, width=38)
            header_entry.insert(0, "1")
            header_entry.grid(row=0, column=3, padx=2)
            header_entry.bind("<FocusOut>", lambda _event, side="source" if column == 0 else "destination": self.reload_reconcile_columns(side))
            if column == 0:
                self.reconcile_source_header_start_entry = header_start_entry
                self.reconcile_source_header_entry = header_entry
            else:
                self.reconcile_destination_header_start_entry = header_start_entry
                self.reconcile_destination_header_entry = header_entry

            ctk.CTkLabel(panel, text="Click để chọn; click lại để bỏ chọn (không cần giữ Ctrl):").grid(row=3, column=0, columnspan=2, pady=(5, 2))
            listbox = tk.Listbox(panel, selectmode=tk.MULTIPLE, exportselection=False, height=8,
                                 bg="#333333", fg="white", selectbackground="#2ecc71")
            listbox.grid(row=4, column=0, columnspan=2, padx=8, pady=(0, 3), sticky="ew")
            listbox.bind("<Control-a>", lambda _event, lb=listbox: (lb.select_set(0, tk.END), "break")[1])
            listbox.bind("<Control-A>", lambda _event, lb=listbox: (lb.select_set(0, tk.END), "break")[1])
            selection_buttons = ctk.CTkFrame(panel, fg_color="transparent")
            selection_buttons.grid(row=5, column=0, columnspan=2, pady=(0, 6))
            ctk.CTkButton(
                selection_buttons, text="Chọn tất cả", width=95,
                command=lambda selected_side=side: self.select_all_reconcile_columns(selected_side),
            ).grid(row=0, column=0, padx=3)
            ctk.CTkButton(
                selection_buttons, text="Bỏ chọn", width=85, fg_color="#7f8c8d",
                command=lambda selected_side=side: self.clear_reconcile_column_selection(selected_side),
            ).grid(row=0, column=1, padx=3)
            if column == 0:
                self.reconcile_source_columns = listbox
            else:
                self.reconcile_destination_columns = listbox

        options = ctk.CTkFrame(self.reconcile_frame, fg_color="transparent")
        options.grid(row=2, column=0, columnspan=2, padx=20, pady=8)
        ctk.CTkLabel(options, text="Tên cột ghi chú:").grid(row=0, column=0, padx=4)
        self.reconcile_note_header_entry = ctk.CTkEntry(options, width=160)
        self.reconcile_note_header_entry.insert(0, "Ghi chú đối soát")
        self.reconcile_note_header_entry.grid(row=0, column=1, padx=4)
        ctk.CTkLabel(options, text="Nội dung khi tìm thấy:").grid(row=0, column=2, padx=(14, 4))
        self.reconcile_note_entry = ctk.CTkEntry(options, width=120)
        self.reconcile_note_entry.insert(0, "Có")
        self.reconcile_note_entry.grid(row=0, column=3, padx=4)
        self.reconcile_ignore_case_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(options, text="Không phân biệt hoa/thường", variable=self.reconcile_ignore_case_var).grid(row=0, column=4, padx=14)
        ctk.CTkLabel(options, text="Lấy thêm dữ liệu cột nguồn:").grid(row=1, column=0, padx=4, pady=(8, 0))
        self.reconcile_fill_column_var = ctk.StringVar(value="Không lấy thêm")
        self.reconcile_fill_column_menu = ctk.CTkOptionMenu(
            options, variable=self.reconcile_fill_column_var, values=["Không lấy thêm"], width=230
        )
        self.reconcile_fill_column_menu.grid(row=1, column=1, columnspan=2, padx=4, pady=(8, 0), sticky="w")
        ctk.CTkLabel(options, text="Tên cột kết quả:").grid(row=1, column=3, padx=4, pady=(8, 0))
        self.reconcile_fill_header_entry = ctk.CTkEntry(options, width=170)
        self.reconcile_fill_header_entry.insert(0, "Dữ liệu từ nguồn")
        self.reconcile_fill_header_entry.grid(row=1, column=4, padx=4, pady=(8, 0))

        reconcile_actions = ctk.CTkFrame(self.reconcile_frame, fg_color="transparent")
        reconcile_actions.grid(row=3, column=0, columnspan=2, pady=8)
        self.reconcile_pairs_btn = ctk.CTkButton(
            reconcile_actions, text="Danh sách cặp file (0)", command=self.show_reconcile_pairs_window,
            height=40, width=170, fg_color="#2980b9",
        )
        self.reconcile_pairs_btn.grid(row=0, column=0, padx=5)
        self.reconcile_btn = ctk.CTkButton(
            reconcile_actions, text="▶  Bắt đầu Đối Soát", command=self.reconcile_process,
            height=45, fg_color="#16a085", font=ctk.CTkFont(size=14, weight="bold"),
        )
        self.reconcile_btn.grid(row=0, column=1, padx=5)
        self.reconcile_progress_frame, self.reconcile_progress_bar, self.reconcile_status_label = self._create_progress_widgets(
            self.reconcile_frame, row=4, columnspan=2
        )

    def setup_guide_frame(self):
        ctk.CTkLabel(
            self.guide_frame,
            text="📖  Hướng Dẫn Sử Dụng",
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(pady=20)
        
        textbox = ctk.CTkTextbox(self.guide_frame, width=800, height=550, font=ctk.CTkFont(size=13))
        textbox.pack(padx=20, pady=10, fill="both", expand=True)
        
        guide_text = """HƯỚNG DẪN SỬ DỤNG EXCEL TOOLBOX

1. GHÉP SHEET TRONG CÙNG MỘT FILE
• Chọn file Excel, sau đó chọn một hoặc nhiều sheet (giữ Ctrl/Shift hoặc dùng nút Chọn tất cả Sheet).
• Khai báo vùng tiêu đề từ dòng–đến dòng và vùng dữ liệu. Dòng kết thúc dữ liệu = 0 nghĩa là tự tìm dòng cuối.
• Các tùy chọn:
  - Dán giá trị: chỉ lấy kết quả hiển thị, không giữ công thức.
  - Đánh lại STT cột A: tạo lại số thứ tự liên tục.
  - Ưu tiên Excel COM: giữ định dạng tốt nhất, yêu cầu máy có Microsoft Excel.
  - Copy nhanh: tăng tốc nhưng không giữ đầy đủ định dạng.
• Bấm Bắt đầu Ghép Sheet. Kết quả được ghi vào sheet TongHop của file mới.

2. GHÉP NHIỀU FILE EXCEL
• Bấm Thêm Files để chọn nhiều file; dùng Xóa để xóa danh sách đã chọn nếu cần.
• Nhập vùng dòng tiêu đề. Dữ liệu bắt đầu từ dòng ngay sau dòng tiêu đề cuối.
• Chọn Giữ tiêu đề của mỗi sheet/file nếu muốn lặp lại tiêu đề cho từng phần.
• Có thể chọn Dán giá trị, Ưu tiên Excel COM hoặc Copy nhanh như mục 1.
• Bấm Bắt đầu Ghép File và chọn nơi lưu kết quả.

3. TÁCH FILE EXCEL
• Chọn file và sheet cần tách, sau đó khai báo vùng tiêu đề.
• Tách theo giá trị cột: bấm Tải lại cột, chọn cột; các dòng cùng giá trị được đưa vào cùng một phần.
• Tách theo số lượng dòng: nhập số dòng tối đa cho mỗi phần.
• Chọn xuất thành nhiều sheet trong một file hoặc nhiều file riêng biệt.
• Copy nhanh giúp xử lý nhanh hơn nhưng không giữ đầy đủ định dạng.
• Bấm Bắt đầu Tách File và chọn nơi lưu.

4. ĐỐI SOÁT DỮ LIỆU THEO NHIỀU CỘT
• Tải NGUỒN ở bên trái (file dùng để tìm/tra cứu) và ĐÍCH ở bên phải (file sẽ nhận cột ghi chú/kết quả), rồi chọn đúng sheet ở mỗi bên.
• Có thể chọn nhiều file cùng lúc. Phần mềm ghép theo thứ tự chọn: Nguồn 1 ↔ Đích 1, Nguồn 2 ↔ Đích 2... Bấm Danh sách cặp file để xem; mỗi cặp nằm trên một dòng. Số file hai bên phải bằng nhau.
• Trong cửa sổ Danh sách cặp file, chọn từng dòng rồi chọn Sheet nguồn và Sheet đích riêng cho cặp đó. Vùng tiêu đề và vị trí cột khóa trên màn hình chính được áp dụng cho mọi cặp; mỗi file đích tạo một file kết quả riêng.
• Nếu một cặp bị lỗi, hộp thoại ghi rõ STT, file nguồn, file đích và nguyên nhân. Chọn Tiếp tục – Bỏ qua cặp này để chạy cặp kế tiếp, hoặc Dừng đối soát.
• Sau mỗi lần chạy, phần mềm tự tạo BaoCao_DoiSoat_*.xlsx; mỗi cặp một dòng với trạng thái Thành công, Lỗi hoặc Chưa thực hiện, số dòng kiểm tra/khớp, file kết quả và chi tiết lỗi.
• Nút Xóa đặt lại file, sheet và danh sách cột của từng bên để nạp workbook mới.
• Khai báo vùng tiêu đề từ dòng–đến dòng. Tên cột được đọc tại dòng tiêu đề cuối và dữ liệu bắt đầu ở dòng kế tiếp.
• Click một lần để chọn cột; click lại chính cột đó để bỏ chọn, không cần giữ Ctrl. Có thể dùng Ctrl + A hoặc nút Chọn tất cả; nút Bỏ chọn xóa toàn bộ vùng chọn. Hai bên phải có cùng số cột và cùng thứ tự ghép.
  Ví dụ: chọn C rồi D ở nguồn và C rồi D ở đích để đối soát theo cặp (C, D).
• Nhập tên cột ghi chú và nội dung khi tìm thấy, ví dụ Ghi chú đối soát / Có.
• Tùy chọn Không phân biệt hoa/thường cũng tự loại khoảng trắng thừa và chuẩn hóa số nguyên (107.0 tương đương 107).
• Nếu cần lấy thêm dữ liệu từ nguồn, chọn cột tại mục Lấy thêm dữ liệu cột nguồn (ví dụ cột T), rồi đặt Tên cột kết quả.
• Bấm Bắt đầu Đối Soát và chọn file kết quả. Phần mềm tạo bản sao file đích, sau đó:
  - Ghi nội dung Có cho dòng có toàn bộ khóa xuất hiện trên cùng một dòng nguồn.
  - Sao chép dữ liệu từ cột nguồn đã chọn vào cột kết quả mới.
  - Dòng không khớp được để trống.
• Thanh tiến độ hiển thị tỷ lệ %, số dòng đã xử lý và số dòng đang khớp. Khi hoàn tất, phần mềm hỏi bạn có muốn mở file kết quả ngay hay không.
• Nếu một khóa nguồn xuất hiện nhiều lần, phần mềm ưu tiên giá trị không trống đầu tiên của cột lấy thêm.
• File kết quả phải khác file đích để bảo vệ dữ liệu gốc.

LƯU Ý CHUNG
• Phiên bản hiện tại được hiển thị ở cuối thanh điều hướng. Ứng dụng tự kiểm tra khi khởi động; có thể bấm Kiểm tra cập nhật để kiểm tra thủ công. Khi có bản mới, phần mềm hiển thị phiên bản hiện tại/mới và hỏi trước khi tải cài đặt.
• Hỗ trợ file .xlsx và tự chuyển .xls qua Excel COM khi máy có Microsoft Excel.
• Nên đóng file Excel đang xử lý để tránh lỗi khóa file.
• Luôn kiểm tra đúng sheet, vùng tiêu đề và thứ tự cột khóa trước khi chạy.
• Với dữ liệu lớn, hãy chờ thông báo hoàn tất và không đóng ứng dụng giữa quá trình."""
        textbox.insert("0.0", guide_text)
        textbox.configure(state="disabled")

    def select_all_home_sheets(self):
        self.home_sheet_listbox.select_set(0, tk.END)

    def _create_progress_widgets(self, parent, row, column=0, columnspan=1):
        """Create progress bar and status label, initially hidden."""
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.grid(row=row, column=column, columnspan=columnspan, padx=20, pady=(0, 10), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_remove()

        status_label = ctk.CTkLabel(frame, text="", font=ctk.CTkFont(size=13))
        status_label.pack(pady=(5, 2))

        progress_bar = ctk.CTkProgressBar(frame, width=400, mode="indeterminate")
        progress_bar.pack(pady=(2, 5))
        progress_bar.set(0)

        return frame, progress_bar, status_label

    def _run_reconcile_in_thread(self, task_func):
        """Run reconciliation with real determinate progress and offer to open the result."""
        self.reconcile_btn.configure(state="disabled")
        self.reconcile_progress_bar.stop()
        self.reconcile_progress_bar.configure(mode="determinate")
        self.reconcile_progress_bar.set(0)
        self.reconcile_status_label.configure(
            text="0%  •  Đang chuẩn bị đối soát...", text_color="#f39c12"
        )
        self.reconcile_progress_frame.grid()

        def progress_callback(progress, status):
            self.after(0, lambda p=progress, s=status: self._update_reconcile_progress(p, s))

        def wrapper():
            try:
                message, saved_path = task_func(progress_callback)
                self.after(0, lambda: self._on_reconcile_done(message, saved_path, None))
            except Exception as exc:
                self.after(0, lambda error=exc: self._on_reconcile_done(None, None, error))

        threading.Thread(target=wrapper, daemon=True).start()

    def _update_reconcile_progress(self, progress, status):
        percent = int(round(progress * 100))
        self.reconcile_progress_bar.set(progress)
        self.reconcile_status_label.configure(
            text=f"{percent}%  •  {status}", text_color="#3498db" if progress < 1 else "#2ecc71"
        )

    def _on_reconcile_done(self, message, saved_path, error):
        self.reconcile_btn.configure(state="normal")
        if error:
            self.reconcile_progress_bar.set(0)
            self.reconcile_status_label.configure(text=f"❌  Lỗi: {error}", text_color="#e74c3c")
            messagebox.showerror("Lỗi đối soát", str(error))
        else:
            self.reconcile_progress_bar.set(1)
            self.reconcile_status_label.configure(
                text="100%  •  Đã đối soát và lưu file thành công", text_color="#2ecc71"
            )
            target_is_folder = os.path.isdir(saved_path)
            target_label = "thư mục chứa các file kết quả" if target_is_folder else "file kết quả vừa tạo"
            should_open = messagebox.askyesno(
                "Đối soát hoàn tất",
                f"{message}\n\nBạn có muốn mở {target_label} không?",
            )
            if should_open:
                try:
                    os.startfile(os.path.abspath(saved_path))
                except Exception as exc:
                    messagebox.showerror("Không thể mở kết quả", f"Kết quả đã được lưu nhưng không thể mở:\n{exc}")

        self.after(15000, lambda: self.reconcile_progress_frame.grid_remove())

    def _run_in_thread(self, task_func, action_button, progress_frame, progress_bar, status_label):
        """Run task_func in a background thread with progress indication."""
        action_button.configure(state="disabled")
        status_label.configure(text="⏳  Đang xử lý, vui lòng chờ...", text_color="#f39c12")
        progress_frame.grid()
        progress_bar.start()

        def wrapper():
            try:
                result_msg = task_func()
                self.after(0, lambda: self._on_task_done(
                    action_button, progress_frame, progress_bar, status_label, result_msg, None
                ))
            except Exception as exc:
                self.after(0, lambda: self._on_task_done(
                    action_button, progress_frame, progress_bar, status_label, None, exc
                ))

        threading.Thread(target=wrapper, daemon=True).start()

    def _on_task_done(self, action_button, progress_frame, progress_bar, status_label, result_msg, error):
        """Handle task completion on the main thread."""
        progress_bar.stop()
        action_button.configure(state="normal")

        if error:
            status_label.configure(text=f"\u274c  Lỗi: {error}", text_color="#e74c3c")
            messagebox.showerror("Lỗi", str(error))
        else:
            status_label.configure(text=f"\u2705  {result_msg}", text_color="#2ecc71")
            messagebox.showinfo("Thành công", result_msg)

        self.after(8000, lambda: progress_frame.grid_remove())

    def toggle_split_inputs(self):
        if self.split_mode_var.get() == "column":
            self.split_header_menu.grid()
            self.row_count_entry.grid_remove()
        else:
            self.split_header_menu.grid_remove()
            self.row_count_entry.grid(row=0, column=4, padx=5)

    def select_frame_by_name(self, name):
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "merge_sheet" else "transparent")
        self.frame_2_button.configure(fg_color=("gray75", "gray25") if name == "merge_files" else "transparent")
        self.frame_3_button.configure(fg_color=("gray75", "gray25") if name == "split_file" else "transparent")
        self.reconcile_button.configure(fg_color=("gray75", "gray25") if name == "reconcile" else "transparent")
        self.frame_4_button.configure(fg_color=("gray75", "gray25") if name == "guide" else "transparent")

        self.home_frame.grid_forget()
        self.second_frame.grid_forget()
        self.third_frame.grid_forget()
        self.reconcile_frame.grid_forget()
        self.guide_frame.grid_forget()
        if name == "merge_sheet":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        elif name == "merge_files":
            self.second_frame.grid(row=0, column=1, sticky="nsew")
        elif name == "split_file":
            self.third_frame.grid(row=0, column=1, sticky="nsew")
        elif name == "reconcile":
            self.reconcile_frame.grid(row=0, column=1, sticky="nsew")
        elif name == "guide":
            self.guide_frame.grid(row=0, column=1, sticky="nsew")

    def browse_reconcile_file(self, side):
        file_paths = list(filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")]))
        if not file_paths:
            return
        file_path = file_paths[0]
        path_var = self.reconcile_source_path if side == "source" else self.reconcile_destination_path
        sheet_var = self.reconcile_source_sheet if side == "source" else self.reconcile_destination_sheet
        sheet_menu = self.reconcile_source_sheet_menu if side == "source" else self.reconcile_destination_sheet_menu
        if side == "source":
            self.reconcile_source_files = file_paths
        else:
            self.reconcile_destination_files = file_paths
        path_var.set(file_path if len(file_paths) == 1 else f"{len(file_paths)} file | {file_path}")
        try:
            sheet_options = [get_workbook_sheet_names(path) for path in file_paths]
            sheet_names = sheet_options[0]
            current_sheet = sheet_var.get()
            selected_sheets = [
                current_sheet if current_sheet in names else (names[0] if names else "Chưa có dữ liệu")
                for names in sheet_options
            ]
            if side == "source":
                self.reconcile_source_sheet_options = sheet_options
                self.reconcile_source_pair_sheets = selected_sheets
            else:
                self.reconcile_destination_sheet_options = sheet_options
                self.reconcile_destination_pair_sheets = selected_sheets
            sheet_menu.configure(values=sheet_names or ["Chưa có dữ liệu"])
            sheet_var.set(selected_sheets[0])
            self.reload_reconcile_columns(side)
            self._refresh_reconcile_pairs_display()
        except Exception as exc:
            messagebox.showerror("Lỗi", str(exc))

    def clear_reconcile_file(self, side):
        """Reset one reconciliation input so another workbook can be loaded cleanly."""
        if side == "source":
            self.reconcile_source_files = []
            self.reconcile_source_pair_sheets = []
            self.reconcile_source_sheet_options = []
            path_var = self.reconcile_source_path
            sheet_var = self.reconcile_source_sheet
            sheet_menu = self.reconcile_source_sheet_menu
            listbox = self.reconcile_source_columns
            start_entry = self.reconcile_source_header_start_entry
            end_entry = self.reconcile_source_header_entry
        else:
            self.reconcile_destination_files = []
            self.reconcile_destination_pair_sheets = []
            self.reconcile_destination_sheet_options = []
            path_var = self.reconcile_destination_path
            sheet_var = self.reconcile_destination_sheet
            sheet_menu = self.reconcile_destination_sheet_menu
            listbox = self.reconcile_destination_columns
            start_entry = self.reconcile_destination_header_start_entry
            end_entry = self.reconcile_destination_header_entry

        path_var.set("")
        sheet_var.set("Chưa có dữ liệu")
        sheet_menu.configure(values=["Chưa có dữ liệu"])
        listbox.delete(0, tk.END)
        for entry in (start_entry, end_entry):
            entry.delete(0, tk.END)
            entry.insert(0, "1")
        if side == "source":
            self.reconcile_fill_column_var.set("Không lấy thêm")
            self.reconcile_fill_column_menu.configure(values=["Không lấy thêm"])
        self._refresh_reconcile_pairs_display()

    def select_all_reconcile_columns(self, side):
        listbox = self.reconcile_source_columns if side == "source" else self.reconcile_destination_columns
        if listbox.size():
            listbox.select_set(0, tk.END)
            listbox.focus_set()

    def clear_reconcile_column_selection(self, side):
        listbox = self.reconcile_source_columns if side == "source" else self.reconcile_destination_columns
        listbox.selection_clear(0, tk.END)
        listbox.focus_set()

    def _get_reconcile_pairs(self):
        pair_count = max(len(self.reconcile_source_files), len(self.reconcile_destination_files))
        return [
            (
                self.reconcile_source_files[index] if index < len(self.reconcile_source_files) else None,
                self.reconcile_destination_files[index] if index < len(self.reconcile_destination_files) else None,
            )
            for index in range(pair_count)
        ]

    def show_reconcile_pairs_window(self):
        if self.reconcile_pairs_window is not None and self.reconcile_pairs_window.winfo_exists():
            self.reconcile_pairs_window.lift()
            self._refresh_reconcile_pairs_display()
            return

        window = ctk.CTkToplevel(self)
        window.title("Danh sách cặp file đối soát")
        window.geometry("1180x520")
        window.transient(self)
        window.grid_rowconfigure(1, weight=1)
        window.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            window,
            text="Ghép theo thứ tự: Nguồn 1 ↔ Đích 1, Nguồn 2 ↔ Đích 2, ...",
            font=ctk.CTkFont(size=15, weight="bold"),
        ).grid(row=0, column=0, padx=15, pady=12)

        tree = ttk.Treeview(
            window,
            columns=("order", "source", "source_sheet", "destination", "destination_sheet", "status"),
            show="headings",
        )
        tree.heading("order", text="STT")
        tree.heading("source", text="File nguồn")
        tree.heading("source_sheet", text="Sheet nguồn")
        tree.heading("destination", text="File đích")
        tree.heading("destination_sheet", text="Sheet đích")
        tree.heading("status", text="Trạng thái")
        tree.column("order", width=55, anchor="center", stretch=False)
        tree.column("source", width=270)
        tree.column("source_sheet", width=170)
        tree.column("destination", width=270)
        tree.column("destination_sheet", width=170)
        tree.column("status", width=130, anchor="center")
        tree.grid(row=1, column=0, padx=(15, 0), pady=5, sticky="nsew")
        scrollbar = ttk.Scrollbar(window, orient="vertical", command=tree.yview)
        scrollbar.grid(row=1, column=1, padx=(0, 15), pady=5, sticky="ns")
        tree.configure(yscrollcommand=scrollbar.set)
        tree.bind("<<TreeviewSelect>>", self._on_reconcile_pair_selected)

        sheet_editor = ctk.CTkFrame(window)
        sheet_editor.grid(row=2, column=0, columnspan=2, padx=15, pady=8, sticky="ew")
        ctk.CTkLabel(sheet_editor, text="Chọn dòng ở bảng, sau đó chọn sheet riêng cho cặp:").grid(
            row=0, column=0, columnspan=4, pady=(6, 4)
        )
        ctk.CTkLabel(sheet_editor, text="Sheet nguồn:").grid(row=1, column=0, padx=(12, 4), pady=8)
        self.reconcile_pair_source_sheet_var = ctk.StringVar(value="Chưa chọn cặp")
        self.reconcile_pair_source_sheet_menu = ctk.CTkOptionMenu(
            sheet_editor, variable=self.reconcile_pair_source_sheet_var,
            values=["Chưa chọn cặp"], width=280,
            command=lambda value: self._update_selected_pair_sheet("source", value),
        )
        self.reconcile_pair_source_sheet_menu.grid(row=1, column=1, padx=4, pady=8)
        ctk.CTkLabel(sheet_editor, text="Sheet đích:").grid(row=1, column=2, padx=(18, 4), pady=8)
        self.reconcile_pair_destination_sheet_var = ctk.StringVar(value="Chưa chọn cặp")
        self.reconcile_pair_destination_sheet_menu = ctk.CTkOptionMenu(
            sheet_editor, variable=self.reconcile_pair_destination_sheet_var,
            values=["Chưa chọn cặp"], width=280,
            command=lambda value: self._update_selected_pair_sheet("destination", value),
        )
        self.reconcile_pair_destination_sheet_menu.grid(row=1, column=3, padx=4, pady=8)

        ctk.CTkButton(window, text="Đóng", command=window.destroy, width=100).grid(row=3, column=0, pady=12)
        self.reconcile_pairs_window = window
        self.reconcile_pairs_tree = tree
        window.protocol("WM_DELETE_WINDOW", window.destroy)
        self._refresh_reconcile_pairs_display()

    def _on_reconcile_pair_selected(self, _event=None):
        tree = self.reconcile_pairs_tree
        if tree is None or not tree.winfo_exists() or not tree.selection():
            return
        pair_index = int(tree.selection()[0].split("_")[1]) - 1
        if pair_index < len(self.reconcile_source_sheet_options):
            source_options = self.reconcile_source_sheet_options[pair_index] or ["Chưa có dữ liệu"]
            self.reconcile_pair_source_sheet_menu.configure(values=source_options)
            self.reconcile_pair_source_sheet_var.set(self.reconcile_source_pair_sheets[pair_index])
        if pair_index < len(self.reconcile_destination_sheet_options):
            destination_options = self.reconcile_destination_sheet_options[pair_index] or ["Chưa có dữ liệu"]
            self.reconcile_pair_destination_sheet_menu.configure(values=destination_options)
            self.reconcile_pair_destination_sheet_var.set(self.reconcile_destination_pair_sheets[pair_index])

    def _update_selected_pair_sheet(self, side, value):
        tree = self.reconcile_pairs_tree
        if tree is None or not tree.winfo_exists() or not tree.selection():
            return
        pair_index = int(tree.selection()[0].split("_")[1]) - 1
        selections = (
            self.reconcile_source_pair_sheets if side == "source"
            else self.reconcile_destination_pair_sheets
        )
        if pair_index < len(selections):
            selections[pair_index] = value
            if pair_index == 0:
                if side == "source":
                    self.reconcile_source_sheet.set(value)
                else:
                    self.reconcile_destination_sheet.set(value)
                self.reload_reconcile_columns(side)
            self._refresh_reconcile_pairs_display()

    def _refresh_reconcile_pairs_display(self):
        pairs = self._get_reconcile_pairs()
        complete_count = sum(bool(source and destination) for source, destination in pairs)
        if hasattr(self, "reconcile_pairs_btn"):
            self.reconcile_pairs_btn.configure(text=f"Danh sách cặp file ({complete_count})")
        tree = self.reconcile_pairs_tree
        if tree is None or not tree.winfo_exists():
            return
        selected_item = tree.selection()[0] if tree.selection() else None
        tree.delete(*tree.get_children())
        for index, (source, destination) in enumerate(pairs, start=1):
            status = "Sẵn sàng" if source and destination else "Thiếu file"
            source_sheet = (
                self.reconcile_source_pair_sheets[index - 1]
                if index - 1 < len(self.reconcile_source_pair_sheets) else "—"
            )
            destination_sheet = (
                self.reconcile_destination_pair_sheets[index - 1]
                if index - 1 < len(self.reconcile_destination_pair_sheets) else "—"
            )
            tree.insert(
                "", "end", iid=f"pair_{index}",
                values=(
                    index, os.path.basename(source) if source else "—", source_sheet,
                    os.path.basename(destination) if destination else "—", destination_sheet, status,
                ),
            )
        if selected_item and tree.exists(selected_item):
            tree.selection_set(selected_item)

    def _set_reconcile_pair_status(self, pair_number, status):
        tree = self.reconcile_pairs_tree
        if tree is None or not tree.winfo_exists():
            return
        item_id = f"pair_{pair_number}"
        if tree.exists(item_id):
            values = list(tree.item(item_id, "values"))
            values[5] = status
            tree.item(item_id, values=values)
            tree.see(item_id)

    def _ask_continue_after_pair_error(self, pair_number, source_path, destination_path, error_text):
        """Show a main-thread error dialog and wait for Continue or Stop from the worker."""
        completed = threading.Event()
        decision = {"continue": False}

        def show_dialog():
            dialog = ctk.CTkToplevel(self)
            dialog.title(f"Lỗi đối soát tại cặp {pair_number}")
            dialog.geometry("760x390")
            dialog.transient(self)
            dialog.grab_set()
            dialog.grid_columnconfigure(0, weight=1)
            dialog.grid_rowconfigure(2, weight=1)
            ctk.CTkLabel(
                dialog, text=f"❌  Cặp đối soát số {pair_number} bị lỗi",
                font=ctk.CTkFont(size=18, weight="bold"), text_color="#e74c3c",
            ).grid(row=0, column=0, padx=20, pady=(18, 8))
            ctk.CTkLabel(
                dialog,
                text=(f"Nguồn: {os.path.basename(source_path)}\n"
                      f"Đích: {os.path.basename(destination_path)}"),
                justify="left", anchor="w",
            ).grid(row=1, column=0, padx=25, pady=5, sticky="ew")
            error_box = ctk.CTkTextbox(dialog, height=150, wrap="word")
            error_box.grid(row=2, column=0, padx=25, pady=8, sticky="nsew")
            error_box.insert("0.0", error_text)
            error_box.configure(state="disabled")
            buttons = ctk.CTkFrame(dialog, fg_color="transparent")
            buttons.grid(row=3, column=0, pady=15)

            def close_dialog(should_continue):
                decision["continue"] = should_continue
                try:
                    dialog.grab_release()
                except Exception:
                    pass
                dialog.destroy()
                completed.set()

            ctk.CTkButton(
                buttons, text="Tiếp tục – Bỏ qua cặp này", width=210,
                fg_color="#16a085", command=lambda: close_dialog(True),
            ).grid(row=0, column=0, padx=8)
            ctk.CTkButton(
                buttons, text="Dừng đối soát", width=150,
                fg_color="#c0392b", command=lambda: close_dialog(False),
            ).grid(row=0, column=1, padx=8)
            dialog.protocol("WM_DELETE_WINDOW", lambda: close_dialog(False))
            dialog.lift()

        self.after(0, show_dialog)
        completed.wait()
        return decision["continue"]

    def on_reconcile_main_sheet_changed(self, side, value):
        """Keep the first pair synchronized with the main sheet selector."""
        selections = (
            self.reconcile_source_pair_sheets if side == "source"
            else self.reconcile_destination_pair_sheets
        )
        if selections:
            selections[0] = value
        self.reload_reconcile_columns(side)
        self._refresh_reconcile_pairs_display()

    def reload_reconcile_columns(self, side):
        files = self.reconcile_source_files if side == "source" else self.reconcile_destination_files
        path = files[0] if files else ""
        sheet = self.reconcile_source_sheet.get() if side == "source" else self.reconcile_destination_sheet.get()
        entry = self.reconcile_source_header_entry if side == "source" else self.reconcile_destination_header_entry
        listbox = self.reconcile_source_columns if side == "source" else self.reconcile_destination_columns
        if not path or not os.path.exists(path) or sheet == "Chưa có dữ liệu":
            return
        try:
            header_row = int(entry.get())
            columns = get_header_columns(path, sheet, header_row)
            listbox.delete(0, tk.END)
            for value in columns:
                listbox.insert(tk.END, value)
            if side == "source":
                fill_values = ["Không lấy thêm"] + columns
                self.reconcile_fill_column_menu.configure(values=fill_values)
                if self.reconcile_fill_column_var.get() not in fill_values:
                    self.reconcile_fill_column_var.set("Không lấy thêm")
        except Exception as exc:
            messagebox.showerror("Lỗi", f"Không thể tải danh sách cột: {exc}")

    def reconcile_process(self):
        pairs = self._get_reconcile_pairs()
        if not pairs:
            messagebox.showwarning("Cảnh báo", "Vui lòng tải cả danh sách nguồn và danh sách đích.")
            return
        if any(not source or not destination for source, destination in pairs):
            messagebox.showwarning(
                "Cảnh báo",
                "Số file nguồn và file đích phải bằng nhau. Mỗi dòng cần đủ một file nguồn và một file đích.",
            )
            self.show_reconcile_pairs_window()
            return
        source_path, destination_path = pairs[0]
        source_selected = self.reconcile_source_columns.curselection()
        destination_selected = self.reconcile_destination_columns.curselection()
        if not source_selected or not destination_selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn các cột khóa ở cả hai danh sách.")
            return
        if len(source_selected) != len(destination_selected):
            messagebox.showwarning("Cảnh báo", "Hai bên phải chọn cùng số lượng cột khóa và theo cùng thứ tự.")
            return
        note = self.reconcile_note_entry.get()
        if not note:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập nội dung ghi chú khi tìm thấy.")
            return
        try:
            source_header_start = int(self.reconcile_source_header_start_entry.get())
            source_header = int(self.reconcile_source_header_entry.get())
            destination_header_start = int(self.reconcile_destination_header_start_entry.get())
            destination_header = int(self.reconcile_destination_header_entry.get())
            if (source_header_start < 1 or source_header < source_header_start or
                    destination_header_start < 1 or destination_header < destination_header_start):
                raise ValueError
        except ValueError:
            messagebox.showwarning("Cảnh báo", "Dải dòng tiêu đề không hợp lệ (dòng đến phải lớn hơn hoặc bằng dòng từ).")
            return

        if len(pairs) == 1:
            output_target = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{os.path.splitext(os.path.basename(destination_path))[0]}_DoiSoat.xlsx",
            )
        else:
            output_target = filedialog.askdirectory(title="Chọn thư mục lưu các file kết quả đối soát")
        if not output_target:
            return
        source_columns = [get_column_index(self.reconcile_source_columns.get(i).split(" | ")[0]) + 1 for i in source_selected]
        destination_columns = [get_column_index(self.reconcile_destination_columns.get(i).split(" | ")[0]) + 1 for i in destination_selected]
        fill_selection = self.reconcile_fill_column_var.get()
        source_fill_column = None
        if fill_selection and fill_selection != "Không lấy thêm":
            source_fill_column = get_column_index(fill_selection.split(" | ")[0]) + 1
        # Snapshot all Tk values on the main thread before starting background work.
        source_sheet = self.reconcile_source_sheet.get()
        destination_sheet = self.reconcile_destination_sheet.get()
        note_header = self.reconcile_note_header_entry.get()
        fill_header = self.reconcile_fill_header_entry.get()
        ignore_case = self.reconcile_ignore_case_var.get()

        def task(progress_callback):
            total_matched = 0
            total_checked = 0
            total_source_keys = 0
            saved_paths = []
            report_rows = []
            stopped_early = False
            pair_count = len(pairs)
            for pair_index, (pair_source, pair_destination) in enumerate(pairs):
                pair_number = pair_index + 1
                pair_source_sheet = self.reconcile_source_pair_sheets[pair_index]
                pair_destination_sheet = self.reconcile_destination_pair_sheets[pair_index]
                self.after(0, lambda n=pair_number: self._set_reconcile_pair_status(n, "Đang xử lý"))
                try:
                    source_sheets = get_workbook_sheet_names(pair_source)
                    destination_sheets = get_workbook_sheet_names(pair_destination)
                    if pair_source_sheet not in source_sheets:
                        raise ValueError(
                            f"File nguồn không có sheet '{pair_source_sheet}'."
                        )
                    if pair_destination_sheet not in destination_sheets:
                        raise ValueError(
                            f"File đích không có sheet '{pair_destination_sheet}'."
                        )

                    if pair_count == 1:
                        pair_output = output_target
                    else:
                        base_name = os.path.splitext(os.path.basename(pair_destination))[0]
                        pair_output = os.path.join(output_target, f"{base_name}_DoiSoat.xlsx")
                        suffix = 2
                        while os.path.exists(pair_output) or pair_output in saved_paths:
                            pair_output = os.path.join(output_target, f"{base_name}_DoiSoat_{suffix}.xlsx")
                            suffix += 1

                    def pair_progress(progress, status, index=pair_index):
                        overall = (index + progress) / pair_count
                        progress_callback(overall, f"Cặp {index + 1}/{pair_count} • {status}")

                    matched, checked, source_key_count, saved_path = reconcile_excel_files(
                        pair_source, pair_destination, pair_output,
                        pair_source_sheet, pair_destination_sheet,
                        source_columns, destination_columns, source_header, destination_header,
                        note_header, note, ignore_case,
                        source_fill_column, fill_header, pair_progress,
                    )
                    total_matched += matched
                    total_checked += checked
                    total_source_keys += source_key_count
                    saved_paths.append(saved_path)
                    report_rows.append({
                        "index": pair_number, "source": pair_source, "destination": pair_destination,
                        "source_sheet": pair_source_sheet, "destination_sheet": pair_destination_sheet,
                        "status": "Thành công", "checked": checked, "matched": matched,
                        "source_keys": source_key_count, "output": saved_path, "error": "",
                    })
                    self.after(
                        0, lambda n=pair_number, m=matched: self._set_reconcile_pair_status(n, f"Xong • {m:,} khớp")
                    )
                except Exception as exc:
                    error_text = str(exc)
                    report_rows.append({
                        "index": pair_number, "source": pair_source, "destination": pair_destination,
                        "source_sheet": pair_source_sheet, "destination_sheet": pair_destination_sheet,
                        "status": "Lỗi - đã bỏ qua", "error": error_text,
                    })
                    self.after(0, lambda n=pair_number: self._set_reconcile_pair_status(n, "Lỗi"))
                    should_continue = self._ask_continue_after_pair_error(
                        pair_number, pair_source, pair_destination, error_text
                    )
                    if not should_continue:
                        report_rows[-1]["status"] = "Lỗi - người dùng dừng"
                        stopped_early = True
                        for remaining_index in range(pair_index + 1, pair_count):
                            remaining_source, remaining_destination = pairs[remaining_index]
                            report_rows.append({
                                "index": remaining_index + 1, "source": remaining_source,
                                "destination": remaining_destination,
                                "source_sheet": self.reconcile_source_pair_sheets[remaining_index],
                                "destination_sheet": self.reconcile_destination_pair_sheets[remaining_index],
                                "status": "Chưa thực hiện",
                                "error": "Dừng theo lựa chọn của người dùng.",
                            })
                            self.after(
                                0, lambda n=remaining_index + 1: self._set_reconcile_pair_status(n, "Chưa thực hiện")
                            )
                        break

            report_directory = (
                os.path.dirname(os.path.abspath(output_target)) if pair_count == 1 else os.path.abspath(output_target)
            )
            report_path = write_reconcile_batch_report(report_directory, report_rows)
            success_count = sum(row.get("status") == "Thành công" for row in report_rows)
            error_count = sum(str(row.get("status", "")).startswith("Lỗi") for row in report_rows)
            if pair_count == 1 and saved_paths and not error_count:
                message = format_reconcile_result_message(
                    source_columns, destination_columns, total_checked, total_matched,
                    total_source_keys, saved_paths[0]
                )
                open_target = saved_paths[0]
            else:
                message = (
                    f"Đã kết thúc xử lý {pair_count} cặp file.\n"
                    f"Thành công: {success_count} cặp • Lỗi: {error_count} cặp"
                    f"{' • Đã dừng sớm' if stopped_early else ''}.\n"
                    f"Tổng số dòng đích đã kiểm tra: {total_checked:,}.\n"
                    f"Tổng số dòng khớp: {total_matched:,}.\n"
                    f"Báo cáo thực hiện: {report_path}"
                )
                open_target = report_directory
            if pair_count == 1 and saved_paths and not error_count:
                message += f"\nBáo cáo thực hiện: {report_path}"
            return message, open_target

        self._run_reconcile_in_thread(task)

    def browse_single_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        self.input_file_path.set(file_path)
        try:
            sheet_names = get_workbook_sheet_names(file_path)
            self.home_sheet_listbox.delete(0, tk.END)
            for name in sheet_names:
                self.home_sheet_listbox.insert(tk.END, name)
            self.home_sheet_listbox.select_set(0, tk.END)
        except Exception as exc:
            messagebox.showerror("Lỗi", str(exc))

    def browse_split_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        self.split_input_path.set(file_path)
        try:
            sheet_names = get_workbook_sheet_names(file_path)
            self.split_sheet_menu.configure(values=sheet_names)
            if sheet_names:
                self.split_sheet_var.set(sheet_names[0])
            self.reload_columns()
        except Exception as exc:
            messagebox.showerror("Lỗi", str(exc))

    def reload_columns(self):
        file_path = self.split_input_path.get()
        if not file_path or not os.path.exists(file_path):
            return
        try:
            header_end = int(self.header_end_entry.get())
            selected_sheet = self.split_sheet_var.get()
            sheet_name = selected_sheet if selected_sheet != "Active Sheet" else 0
            values = get_header_columns(file_path, sheet_name, header_end)
            self.split_header_menu.configure(values=values or ["Chưa có dữ liệu"])
            if values:
                self.split_header_menu.set(values[0])
        except Exception as exc:
            messagebox.showerror("Lỗi", str(exc))

    def split_process(self):
        input_file = self.split_input_path.get()
        if not input_file:
            messagebox.showwarning("Canh bao", "Vui long chon file can tach!")
            return

        try:
            header_start = int(self.header_start_entry.get())
            header_end = int(self.header_end_entry.get())
            if header_start < 1 or header_end < header_start:
                messagebox.showwarning("Cảnh báo", "Dải dòng tiêu đề không hợp lệ!")
                return

            mode = self.split_mode_var.get()
            output_mode = self.split_output_var.get()
            selected_sheet = self.split_sheet_var.get()
            row_count = None
            column_index = None

            if mode == "column":
                selected_value = self.split_header_menu.get()
                if selected_value == "Chưa có dữ liệu" or not selected_value:
                    messagebox.showwarning("Cảnh báo", "Vui lòng tải lại và chọn cột để tách!")
                    return
                column_index = get_column_index(selected_value.split(" | ")[0])
            else:
                row_count = int(self.row_count_entry.get())
                if row_count <= 0:
                    raise ValueError

            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if not save_path:
                return

            prefer_com = getattr(self, "split_prefer_com_var", ctk.BooleanVar(value=True)).get()
            fast_copy = getattr(self, "split_fast_copy_var", ctk.BooleanVar(value=False)).get()
            def task():
                num_parts = split_excel(
                    input_file, save_path, mode, output_mode,
                    header_start, header_end,
                    sheet_name=selected_sheet,
                    column_index=column_index, row_count=row_count, prefer_excel_com=prefer_com,
                    fast_copy=fast_copy,
                )
                return f"Đã tách xong thành {num_parts} phần!"

            self._run_in_thread(task, self.split_btn, self.split_progress_frame,
                                self.split_progress_bar, self.split_status_label)
        except Exception as exc:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {exc}")

    def merge_sheets_process(self):
        input_path = self.input_file_path.get()
        if not input_path:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file!")
            return
        selected_indices = self.home_sheet_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất 1 sheet để ghép!")
            return

        selected_sheets = [self.home_sheet_listbox.get(i) for i in selected_indices]
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:
            return

        try:
            header_start = int(self.merge_header_start_entry.get())
            header_end = int(self.merge_header_end_entry.get())
            data_start = int(self.merge_data_start_entry.get())
            data_end = int(self.merge_data_end_entry.get())
            if header_start < 1 or header_end < header_start:
                messagebox.showwarning("Cảnh báo", "Dải dòng tiêu đề không hợp lệ!")
                return
            if data_start < 1 or (data_end != 0 and data_end < data_start):
                messagebox.showwarning("Cảnh báo", "Dải dòng dữ liệu không hợp lệ! Nhập 0 ở dòng kết thúc để tự động.")
                return
        except ValueError:
            messagebox.showwarning("Cảnh báo", "Giá trị dòng phải là số nguyên!")
            return

        paste_values = self.home_paste_values_var.get()
        prefer_com = self.home_prefer_com_var.get()
        renumber = self.home_renumber_var.get()
        fast_copy = self.home_fast_copy_var.get()

        def task():
            count = merge_selected_sheets_to_summary(
                input_path, output_path, selected_sheets,
                header_start, header_end, data_start, data_end,
                paste_values=paste_values, prefer_excel_com=prefer_com,
                renumber_first_column=renumber, summary_sheet_name="TongHop",
                fast_copy=fast_copy,
            )
            return f"Đã ghép xong {count} sheet vào Tổng Hợp!"

        self._run_in_thread(task, self.home_merge_btn, self.home_progress_frame,
                            self.home_progress_bar, self.home_status_label)

    def browse_multiple_files(self):
        for file_path in filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")]):
            if file_path not in self.file_list:
                self.file_list.append(file_path)
                self.file_listbox.insert(tk.END, os.path.basename(file_path))

    def clear_file_list(self):
        self.file_list = []
        self.file_listbox.delete(0, tk.END)

    def merge_files_process(self):
        if not self.file_list:
            messagebox.showwarning("Cảnh báo", "Vui lòng thêm file cần ghép!")
            return

        try:
            header_start = int(self.files_header_start_entry.get())
            header_end = int(self.files_header_end_entry.get())
            if header_start < 1 or header_end < header_start:
                messagebox.showwarning("Cảnh báo", "Dải dòng tiêu đề không hợp lệ!")
                return
        except ValueError:
            messagebox.showwarning("Cảnh báo", "Dòng tiêu đề phải là số nguyên!")
            return

        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:
            return

        keep_headers = self.files_keep_headers_var.get()
        paste_values = self.files_paste_values_var.get()
        prefer_com = self.files_prefer_com_var.get()
        fast_copy = self.files_fast_copy_var.get()
        files = list(self.file_list)

        def task():
            count = merge_multiple_files(
                files, output_path,
                keep_all_headers=keep_headers,
                paste_values=paste_values,
                prefer_excel_com=prefer_com,
                header_start=header_start,
                header_end=header_end,
                fast_copy=fast_copy,
            )
            return f"Đã ghép xong {count} file!"

        self._run_in_thread(task, self.files_merge_btn, self.files_progress_frame,
                            self.files_progress_bar, self.files_status_label)


if __name__ == "__main__":
    app = ExcelMergerApp()
    app.mainloop()
