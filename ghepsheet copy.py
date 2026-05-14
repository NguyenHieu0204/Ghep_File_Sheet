import tkinter as tk
import customtkinter as ctk
import pandas as pd
from tkinter import filedialog, messagebox
import os
import tempfile
import win32com.client
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter as openpyxl_get_col_letter
from copy import copy

# --- SETTINGS ---
ctk.set_appearance_mode("System")  # Use system light/dark mode
ctk.set_default_color_theme("blue")

# --- HELPER FUNCTIONS ---
def get_column_letter(n):
    """Convert a zero-indexed column number to Excel-style letter (0 -> A, 25 -> Z)."""
    result = ""
    while n >= 0:
        result = chr(65 + (n % 26)) + result
        n = (n // 26) - 1
    return result

def get_column_index(letter):
    """Convert Excel-style letter to zero-indexed column number (A -> 0)."""
    n = 0
    for char in letter.upper():
        n = n * 26 + (ord(char) - 64)
    return n - 1

def ensure_xlsx(file_path):
    """Chuyển đổi .xls sang .xlsx bằng win32com (nếu cần)."""
    if str(file_path).lower().endswith('.xls'):
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            abs_path = os.path.abspath(file_path)
            wb = excel.Workbooks.Open(abs_path)
            temp_dir = tempfile.gettempdir()
            base_name = os.path.basename(file_path)
            xlsx_path = os.path.join(temp_dir, base_name + "x")
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)
            wb.SaveAs(xlsx_path, FileFormat=51)
            wb.Close()
            excel.Quit()
            return xlsx_path
        except Exception as e:
            raise Exception(f"Không thể tự động chuyển đổi file .xls sang .xlsx: {str(e)}\nVui lòng mở file bằng Excel và lưu lại dưới dạng .xlsx.")
    return file_path

def copy_cell_style(source_cell, target_cell):
    """Copies style from source_cell to target_cell."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def get_real_max_row(ws):
    """Lấy số dòng thực sự có chứa dữ liệu thay vì max_row (để tránh trường hợp người dùng tô màu cả cột xuống 1 triệu dòng)."""
    real_max = 0
    # 1. Kiểm tra các ô đã được merge (vì ô merge chỉ chứa dữ liệu ở ô đầu tiên)
    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_row > real_max:
            real_max = merged_range.max_row
            
    # 2. Quét nhanh tất cả các cell thực sự tồn tại trong bộ nhớ
    if hasattr(ws, '_cells'):
        for (r, c), cell in ws._cells.items():
            if cell.value is not None and str(cell.value).strip() != "":
                if r > real_max:
                    real_max = r
    else:
        real_max = ws.max_row
        
    return max(1, real_max)

def copy_sheet_data(source_ws, target_ws, start_row, include_header=True, header_row_count=1):
    row_offset = start_row - 1
    for col_letter, col_dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dimension.width
    source_start_row = 1 if include_header else (header_row_count + 1)
    
    for row_idx, row_dimension in source_ws.row_dimensions.items():
        if row_idx >= source_start_row:
            new_row_idx = row_idx + row_offset - (0 if include_header else header_row_count)
            target_ws.row_dimensions[new_row_idx].height = row_dimension.height

    real_max_row = get_real_max_row(source_ws)
    for row in source_ws.iter_rows(min_row=source_start_row, max_row=real_max_row):
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue # Bỏ qua các ô hoàn toàn trống và không có style để tối ưu hiệu suất
            new_cell = target_ws.cell(row=cell.row + row_offset - (0 if include_header else header_row_count), 
                                      column=cell.column, value=cell.value)
            copy_cell_style(cell, new_cell)
    for merged_range in source_ws.merged_cells.ranges:
        if merged_range.min_row >= source_start_row:
            new_min_row = merged_range.min_row + row_offset - (0 if include_header else header_row_count)
            new_max_row = merged_range.max_row + row_offset - (0 if include_header else header_row_count)
            new_range = f"{openpyxl_get_col_letter(merged_range.min_col)}{new_min_row}:" \
                        f"{openpyxl_get_col_letter(merged_range.max_col)}{new_max_row}"
            try: target_ws.merge_cells(new_range)
            except Exception: pass
    return get_real_max_row(target_ws)

# --- CORE LOGIC ---
def merge_all_sheets(input_path, output_path):
    converted_path = ensure_xlsx(input_path)
    src_wb = load_workbook(converted_path, data_only=False)
    
    # Sử dụng sheet đầu tiên làm base để giữ nguyên hoàn toàn định dạng của workbook
    dst_wb = src_wb
    if not dst_wb.sheetnames:
        return 0
    
    dst_ws = dst_wb.worksheets[0]
    dst_ws.title = "Merged_Sheets"
    
    current_row = get_real_max_row(dst_ws)
    sheet_names = dst_wb.sheetnames.copy()
    
    for i in range(1, len(sheet_names)):
        sheet_name = sheet_names[i]
        ws = dst_wb[sheet_name]
        # Giữ nguyên include_header=True để không làm mất dòng đầu tiên của các sheet (trong trường hợp file không có tiêu đề)
        current_row = copy_sheet_data(ws, dst_ws, current_row + 1, include_header=True)
        
    # Xóa các sheet cũ đi sau khi đã gộp xong
    for sheet_name in sheet_names[1:]:
        del dst_wb[sheet_name]
        
    dst_wb.save(output_path)
    return len(sheet_names)

def merge_multiple_files(file_list, output_path):
    if not file_list: return 0
    
    # Lấy file đầu tiên làm base workbook
    converted_path_0 = ensure_xlsx(file_list[0])
    dst_wb = load_workbook(converted_path_0, data_only=False)
    dst_ws = dst_wb.worksheets[0]
    dst_ws.title = "Merged_Files"
    
    current_row = get_real_max_row(dst_ws)
    sheet_names_0 = dst_wb.sheetnames.copy()
    
    # Gộp các sheet còn lại trong file đầu tiên
    for j in range(1, len(sheet_names_0)):
        ws = dst_wb[sheet_names_0[j]]
        current_row = copy_sheet_data(ws, dst_ws, current_row + 1, include_header=True)
    
    # Xóa các sheet thừa trong file đầu tiên
    for sheet_name in sheet_names_0[1:]:
        del dst_wb[sheet_name]

    # Gộp các file tiếp theo
    for i in range(1, len(file_list)):
        converted_path = ensure_xlsx(file_list[i])
        src_wb = load_workbook(converted_path, data_only=False)
        for j, sheet_name in enumerate(src_wb.sheetnames):
            ws = src_wb[sheet_name]
            current_row = copy_sheet_data(ws, dst_ws, current_row + 1, include_header=True)
            
    dst_wb.save(output_path)
    return len(file_list)

def split_by_header(input_path, column_index, output_path, header_row=0):
    df = pd.read_excel(input_path, header=header_row, dtype=str)
    col_name = df.columns[column_index]
    unique_values = df[col_name].unique()
    with pd.ExcelWriter(output_path) as writer:
        for val in unique_values:
            s_name = str(val)[:30].strip()
            for char in ['/', '\\', '?', '*', ':', '[', ']']: s_name = s_name.replace(char, '')
            if not s_name: s_name = "Sheet_" + str(val)
            df_slice = df[df[col_name] == val]
            df_slice.to_excel(writer, sheet_name=s_name, index=False)
    return len(unique_values)

# --- UI APPLICATION ---
class ExcelMergerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Excel Tool - Duy Hieu - Professional")
        self.geometry("1100x700")

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # Sidebar
        self.navigation_frame = ctk.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)
        
        # Explicit text color for the sidebar header
        self.label_header = ctk.CTkLabel(self.navigation_frame, text="  EXCEL TOOLBOX", 
                                         font=ctk.CTkFont(size=15, weight="bold"),
                                         text_color=("black", "white"))
        self.label_header.grid(row=0, column=0, padx=20, pady=20)

        # Sidebar buttons with explicit text colors for unselected state
        self.home_button = ctk.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, 
                                         text="Ghép Sheet (1 File)", fg_color="transparent", 
                                         text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                         anchor="w", command=lambda: self.select_frame_by_name("ghép_sheet"))
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.frame_2_button = ctk.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, 
                                            text="Ghép Nhiều File", fg_color="transparent", 
                                            text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                            anchor="w", command=lambda: self.select_frame_by_name("ghép_nhiều_file"))
        self.frame_2_button.grid(row=2, column=0, sticky="ew")

        self.frame_3_button = ctk.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, 
                                            text="Tách File Excel", fg_color="transparent", 
                                            text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                            anchor="w", command=lambda: self.select_frame_by_name("tách_file"))
        self.frame_3_button.grid(row=3, column=0, sticky="ew")

        # Frames
        self.home_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent"); self.setup_home_frame()
        self.second_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent"); self.setup_second_frame()
        self.third_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent"); self.setup_third_frame()
        
        self.select_frame_by_name("ghép_sheet")

    def setup_home_frame(self):
        ctk.CTkLabel(self.home_frame, text="Ghép tất cả Sheet (Giữ nguyên định dạng & Merge)", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, padx=20, pady=20)
        self.input_file_path = ctk.StringVar()
        ctk.CTkEntry(self.home_frame, textvariable=self.input_file_path, width=400).grid(row=1, column=0, padx=20, pady=10)
        ctk.CTkButton(self.home_frame, text="Chọn File", command=self.browse_single_file).grid(row=1, column=1, padx=20, pady=10)
        self.sheet_info_label = ctk.CTkLabel(self.home_frame, text="Chưa chọn file", text_color=("gray40", "gray60"))
        self.sheet_info_label.grid(row=2, column=0, columnspan=2)
        ctk.CTkButton(self.home_frame, text="Bắt đầu Ghép Sheet", command=self.merge_sheets_process, height=50, fg_color="#2ecc71").grid(row=3, column=0, columnspan=2, pady=30)

    def setup_second_frame(self):
        ctk.CTkLabel(self.second_frame, text="Ghép nhiều file Excel (Giữ nguyên định dạng & Merge)", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, padx=20, pady=20)
        self.file_list = []
        self.file_listbox = tk.Listbox(self.second_frame, width=80, height=10, bg="#333333", fg="white")
        self.file_listbox.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        file_ops = ctk.CTkFrame(self.second_frame, fg_color="transparent")
        file_ops.grid(row=1, column=1, padx=20, pady=10, sticky="n")
        ctk.CTkButton(file_ops, text="Thêm Files", command=self.browse_multiple_files).grid(row=0, column=0, pady=5)
        ctk.CTkButton(file_ops, text="Xóa", fg_color="#e74c3c", command=self.clear_file_list).grid(row=1, column=0, pady=5)
        ctk.CTkButton(self.second_frame, text="Bắt đầu Ghép File", command=self.merge_files_process, height=50, fg_color="#3498db").grid(row=2, column=0, columnspan=2, pady=30)

    def setup_third_frame(self):
        ctk.CTkLabel(self.third_frame, text="Tách File Excel thành nhiều Sheet", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, padx=20, pady=20)
        self.split_input_path = ctk.StringVar()
        ctk.CTkEntry(self.third_frame, textvariable=self.split_input_path, width=400).grid(row=1, column=0, padx=20, pady=10)
        ctk.CTkButton(self.third_frame, text="Chọn File", command=self.browse_split_file).grid(row=1, column=1, padx=20, pady=10)
        header_row_frame = ctk.CTkFrame(self.third_frame, fg_color="transparent")
        header_row_frame.grid(row=2, column=0, columnspan=2, pady=5)
        ctk.CTkLabel(header_row_frame, text="Dòng tiêu đề:").grid(row=0, column=0, padx=5)
        self.header_row_entry = ctk.CTkEntry(header_row_frame, width=50); self.header_row_entry.insert(0, "1"); self.header_row_entry.grid(row=0, column=1, padx=5)
        ctk.CTkButton(header_row_frame, text="Tải lại cột", width=80, command=self.reload_columns).grid(row=0, column=2, padx=5)
        self.split_header_menu = ctk.CTkOptionMenu(self.third_frame, values=["Chưa có dữ liệu"], width=400)
        self.split_header_menu.grid(row=4, column=0, columnspan=2, pady=5)
        ctk.CTkButton(self.third_frame, text="Bắt đầu Tách File", command=self.split_process, height=50, fg_color="#9b59b6").grid(row=6, column=0, columnspan=2, pady=40)

    def select_frame_by_name(self, name):
        # Configure button colors with high contrast for selected and unselected states
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "ghép_sheet" else "transparent",
                                   text_color=("black", "white") if name == "ghép_sheet" else ("gray10", "gray90"))
        self.frame_2_button.configure(fg_color=("gray75", "gray25") if name == "ghép_nhiều_file" else "transparent",
                                      text_color=("black", "white") if name == "ghép_nhiều_file" else ("gray10", "gray90"))
        self.frame_3_button.configure(fg_color=("gray100", "gray0") if name == "tách_file" else "transparent", # Use explicit contrast
                                      text_color=("black", "white") if name == "tách_file" else ("gray10", "gray90"))
        
        self.home_frame.grid_forget(); self.second_frame.grid_forget(); self.third_frame.grid_forget()
        if name == "ghép_sheet": self.home_frame.grid(row=0, column=1, sticky="nsew")
        elif name == "ghép_nhiều_file": self.second_frame.grid(row=0, column=1, sticky="nsew")
        elif name == "tách_file": self.third_frame.grid(row=0, column=1, sticky="nsew")

    def browse_single_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if f: 
            self.input_file_path.set(f)
            try:
                sheet_count = len(pd.ExcelFile(f).sheet_names)
                self.sheet_info_label.configure(text=f"Phát hiện {sheet_count} sheet", text_color="green")
            except Exception as e:
                messagebox.showerror("Lỗi", str(e))

    def browse_split_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if f: self.split_input_path.set(f); self.reload_columns()

    def reload_columns(self):
        f = self.split_input_path.get()
        if not f or not os.path.exists(f): return
        try:
            h_row = int(self.header_row_entry.get()) - 1
            df = pd.read_excel(f, header=h_row, nrows=1, dtype=str)
            self.split_header_menu.configure(values=[f"{get_column_letter(i)} | {c}" for i, c in enumerate(df.columns)])
            self.split_header_menu.set(f"{get_column_letter(0)} | {df.columns[0]}")
        except Exception as e: messagebox.showerror("Lỗi", str(e))

    def split_process(self):
        input_path = self.split_input_path.get()
        if not input_path: return
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if output_path:
            try:
                count = split_by_header(input_path, get_column_index(self.split_header_menu.get().split(" | ")[0]), output_path, int(self.header_row_entry.get())-1)
                messagebox.showinfo("Thành công", f"Đã tách thành {count} sheet!")
            except Exception as e: messagebox.showerror("Lỗi", str(e))

    def merge_sheets_process(self):
        input_path = self.input_file_path.get(); output_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if input_path and output_path:
            try:
                count = merge_all_sheets(input_path, output_path)
                messagebox.showinfo("Thành công", f"Đã ghép xong {count} sheet!")
            except Exception as e: messagebox.showerror("Lỗi", str(e))

    def browse_multiple_files(self):
        for f in filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")]):
            if f not in self.file_list: self.file_list.append(f); self.file_listbox.insert(tk.END, os.path.basename(f))

    def clear_file_list(self): self.file_list = []; self.file_listbox.delete(0, tk.END)

    def merge_files_process(self):
        if not self.file_list: return
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if output_path:
            try:
                count = merge_multiple_files(self.file_list, output_path)
                messagebox.showinfo("Thành công", f"Đã ghép xong {count} file!")
            except Exception as e: messagebox.showerror("Lỗi", str(e))

if __name__ == "__main__":
    app = ExcelMergerApp()
    app.mainloop()
