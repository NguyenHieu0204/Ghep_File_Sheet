import os
import tempfile
import tkinter as tk
from copy import copy
from tkinter import filedialog, messagebox

import customtkinter as ctk
import win32com.client
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter as openpyxl_get_col_letter


ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

XL_OPENXML_WORKBOOK = 51
XL_VALUES = -4163
XL_FORMATS = -4122
XL_FORMULAS = -4123
XL_BY_ROWS = 1
XL_BY_COLUMNS = 2
XL_PREVIOUS = 2


def get_column_letter(n):
    """Convert zero-based column index to Excel-style letter."""
    result = ""
    while n >= 0:
        result = chr(65 + (n % 26)) + result
        n = (n // 26) - 1
    return result


def get_column_index(letter):
    """Convert Excel-style letter to zero-based column index."""
    n = 0
    for char in letter.upper():
        n = n * 26 + (ord(char) - 64)
    return n - 1


def normalize_xlsx_output_path(output_path):
    root, ext = os.path.splitext(output_path)
    return output_path if ext.lower() == ".xlsx" else root + ".xlsx"


def ensure_xlsx(file_path):
    """Convert .xls to .xlsx with Excel COM when needed."""
    if not str(file_path).lower().endswith(".xls"):
        return file_path

    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        abs_path = os.path.abspath(file_path)
        wb = excel.Workbooks.Open(abs_path)
        xlsx_path = os.path.join(tempfile.gettempdir(), os.path.basename(file_path) + "x")
        if os.path.exists(xlsx_path):
            os.remove(xlsx_path)
        wb.SaveAs(xlsx_path, FileFormat=XL_OPENXML_WORKBOOK)
        return xlsx_path
    except Exception as exc:
        raise RuntimeError(
            "Khong the tu dong chuyen doi file .xls sang .xlsx. "
            "Vui long mo file bang Excel va luu lai duoi dang .xlsx.\n"
            f"Chi tiet loi: {exc}"
        ) from exc
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def get_workbook_sheet_names(file_path):
    converted_path = ensure_xlsx(file_path)
    wb = load_workbook(converted_path, read_only=True, data_only=False)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def get_header_columns(file_path, sheet_name, header_row):
    converted_path = ensure_xlsx(file_path)
    wb = load_workbook(converted_path, read_only=True, data_only=False)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        values = []
        for index, cell in enumerate(ws[header_row], start=1):
            title = cell.value if cell.value is not None else f"Column {index}"
            values.append(f"{get_column_letter(index - 1)} | {title}")
        return values
    finally:
        wb.close()


def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def get_real_max_row(ws):
    """Return the last row that really has content, not just worksheet formatting."""
    real_max = 0
    for merged_range in ws.merged_cells.ranges:
        real_max = max(real_max, merged_range.max_row)

    if hasattr(ws, "_cells"):
        for (row_idx, _), cell in ws._cells.items():
            if cell.value is not None and str(cell.value).strip() != "":
                real_max = max(real_max, row_idx)
    else:
        real_max = ws.max_row
    return max(1, real_max)


def copy_sheet_data(source_ws, target_ws, start_row, include_header=True, header_row_count=1):
    row_offset = start_row - 1
    source_start_row = 1 if include_header else header_row_count + 1

    for col_letter, col_dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dimension.width
        target_ws.column_dimensions[col_letter].hidden = col_dimension.hidden

    for row_idx, row_dimension in source_ws.row_dimensions.items():
        if row_idx >= source_start_row:
            new_row_idx = row_idx + row_offset - (0 if include_header else header_row_count)
            target_ws.row_dimensions[new_row_idx].height = row_dimension.height
            target_ws.row_dimensions[new_row_idx].hidden = row_dimension.hidden

    real_max_row = get_real_max_row(source_ws)
    for row in source_ws.iter_rows(min_row=source_start_row, max_row=real_max_row):
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            target_row = cell.row + row_offset - (0 if include_header else header_row_count)
            new_cell = target_ws.cell(row=target_row, column=cell.column, value=cell.value)
            copy_cell_style(cell, new_cell)

    for merged_range in source_ws.merged_cells.ranges:
        if merged_range.min_row >= source_start_row:
            new_min_row = merged_range.min_row + row_offset - (0 if include_header else header_row_count)
            new_max_row = merged_range.max_row + row_offset - (0 if include_header else header_row_count)
            new_range = (
                f"{openpyxl_get_col_letter(merged_range.min_col)}{new_min_row}:"
                f"{openpyxl_get_col_letter(merged_range.max_col)}{new_max_row}"
            )
            try:
                target_ws.merge_cells(new_range)
            except Exception:
                pass

    return get_real_max_row(target_ws)


def get_excel_used_bounds(ws):
    last_row = 0
    last_col = 0

    try:
        last_row_cell = ws.Cells.Find(
            What="*",
            After=ws.Cells(1, 1),
            LookIn=XL_FORMULAS,
            SearchOrder=XL_BY_ROWS,
            SearchDirection=XL_PREVIOUS,
            MatchCase=False,
        )
        last_col_cell = ws.Cells.Find(
            What="*",
            After=ws.Cells(1, 1),
            LookIn=XL_FORMULAS,
            SearchOrder=XL_BY_COLUMNS,
            SearchDirection=XL_PREVIOUS,
            MatchCase=False,
        )
        if last_row_cell is not None:
            last_row = max(last_row, last_row_cell.Row)
        if last_col_cell is not None:
            last_col = max(last_col, last_col_cell.Column)
    except Exception:
        pass

    try:
        used = ws.UsedRange
        if used is not None:
            last_row = max(last_row, used.Row + used.Rows.Count - 1)
            last_col = max(last_col, used.Column + used.Columns.Count - 1)
    except Exception:
        pass

    return last_row, last_col


def prepare_destination_workbook(excel, sheet_title):
    wb = excel.Workbooks.Add()
    excel.DisplayAlerts = False
    while wb.Worksheets.Count > 1:
        wb.Worksheets(wb.Worksheets.Count).Delete()
    ws = wb.Worksheets(1)
    ws.Name = sheet_title
    return wb, ws


def copy_excel_range(source_ws, target_ws, source_start_row, target_start_row, paste_values=False):
    last_row, last_col = get_excel_used_bounds(source_ws)
    if last_row < source_start_row or last_col < 1:
        return target_start_row - 1

    source_range = source_ws.Range(source_ws.Cells(source_start_row, 1), source_ws.Cells(last_row, last_col))
    target_cell = target_ws.Cells(target_start_row, 1)

    if paste_values:
        source_range.Copy()
        target_cell.PasteSpecial(Paste=XL_VALUES)
        target_cell.PasteSpecial(Paste=XL_FORMATS)
    else:
        source_range.Copy(Destination=target_cell)

    row_count = last_row - source_start_row + 1
    for col_idx in range(1, last_col + 1):
        target_ws.Columns(col_idx).ColumnWidth = source_ws.Columns(col_idx).ColumnWidth
        target_ws.Columns(col_idx).Hidden = source_ws.Columns(col_idx).Hidden
    for offset in range(row_count):
        target_ws.Rows(target_start_row + offset).RowHeight = source_ws.Rows(source_start_row + offset).RowHeight
        target_ws.Rows(target_start_row + offset).Hidden = source_ws.Rows(source_start_row + offset).Hidden

    return target_start_row + row_count - 1


def copy_excel_fixed_range(source_ws, target_ws, source_start_row, source_end_row, target_start_row, paste_values=False):
    if source_end_row < source_start_row:
        return target_start_row - 1

    _, last_col = get_excel_used_bounds(source_ws)
    if last_col < 1:
        return target_start_row - 1

    source_range = source_ws.Range(source_ws.Cells(source_start_row, 1), source_ws.Cells(source_end_row, last_col))
    target_cell = target_ws.Cells(target_start_row, 1)

    if paste_values:
        source_range.Copy()
        target_cell.PasteSpecial(Paste=XL_VALUES)
        target_cell.PasteSpecial(Paste=XL_FORMATS)
    else:
        source_range.Copy(Destination=target_cell)

    row_count = source_end_row - source_start_row + 1
    for col_idx in range(1, last_col + 1):
        target_ws.Columns(col_idx).ColumnWidth = source_ws.Columns(col_idx).ColumnWidth
        target_ws.Columns(col_idx).Hidden = source_ws.Columns(col_idx).Hidden
    for offset in range(row_count):
        target_ws.Rows(target_start_row + offset).RowHeight = source_ws.Rows(source_start_row + offset).RowHeight
        target_ws.Rows(target_start_row + offset).Hidden = source_ws.Rows(source_start_row + offset).Hidden

    return target_start_row + row_count - 1


def get_excel_last_display_row(source_ws, start_row=1):
    used = source_ws.UsedRange
    last_row = used.Row + used.Rows.Count - 1
    last_col = used.Column + used.Columns.Count - 1

    for row_idx in range(last_row, start_row - 1, -1):
        for col_idx in range(1, last_col + 1):
            text = source_ws.Cells(row_idx, col_idx).Text
            if text is not None and str(text).strip() != "":
                return row_idx
    return start_row - 1


def replace_summary_sheet(workbook, sheet_name):
    excel = workbook.Application
    excel.DisplayAlerts = False
    for idx in range(workbook.Worksheets.Count, 0, -1):
        if workbook.Worksheets(idx).Name == sheet_name:
            workbook.Worksheets(idx).Delete()
            break
    ws = workbook.Worksheets.Add(Before=workbook.Worksheets(1))
    ws.Name = sheet_name
    return ws


def merge_selected_sheets_to_summary_excel_com(
    input_path,
    output_path,
    selected_sheet_names,
    header_start_row,
    header_end_row,
    data_start_row,
    data_end_row=0,
    paste_values=False,
    renumber_first_column=True,
    summary_sheet_name="TongHop",
):
    if not selected_sheet_names:
        return 0

    excel = None
    wb = None
    try:
        converted_path = ensure_xlsx(input_path)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(converted_path), ReadOnly=False)
        summary_ws = replace_summary_sheet(wb, summary_sheet_name)

        first_source_ws = wb.Worksheets(selected_sheet_names[0])
        current_row = 1
        current_row = copy_excel_fixed_range(
            first_source_ws, summary_ws, header_start_row, header_end_row, current_row, paste_values
        )
        current_row += 1

        sequence_number = 1
        for sheet_name in selected_sheet_names:
            source_ws = wb.Worksheets(sheet_name)
            source_end_row = data_end_row if data_end_row and data_end_row >= data_start_row else get_excel_last_display_row(
                source_ws, data_start_row
            )
            if source_end_row < data_start_row:
                continue

            target_start_row = current_row
            current_row = copy_excel_fixed_range(
                source_ws, summary_ws, data_start_row, source_end_row, target_start_row, paste_values
            )

            if renumber_first_column:
                for row_idx in range(target_start_row, current_row + 1):
                    summary_ws.Cells(row_idx, 1).Value = sequence_number
                    sequence_number += 1
            current_row += 1

        output_path = normalize_xlsx_output_path(output_path)
        if os.path.exists(output_path):
            os.remove(output_path)
        wb.SaveAs(os.path.abspath(output_path), FileFormat=XL_OPENXML_WORKBOOK)
        return len(selected_sheet_names)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        if excel is not None:
            try:
                excel.CutCopyMode = False
            except Exception:
                pass
            excel.Quit()


def merge_selected_sheets_excel_com(
    input_path,
    output_path,
    selected_sheet_names,
    keep_all_headers=True,
    paste_values=False,
    header_row_count=1,
):
    if not selected_sheet_names:
        return 0

    excel = None
    src_wb = None
    dst_wb = None
    try:
        converted_path = ensure_xlsx(input_path)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        src_wb = excel.Workbooks.Open(os.path.abspath(converted_path), ReadOnly=True)
        dst_wb, dst_ws = prepare_destination_workbook(excel, "Merged_Sheets")

        current_row = 1
        for index, sheet_name in enumerate(selected_sheet_names):
            source_ws = src_wb.Worksheets(sheet_name)
            include_header = keep_all_headers or index == 0
            source_start_row = 1 if include_header else header_row_count + 1
            current_row = copy_excel_range(source_ws, dst_ws, source_start_row, current_row, paste_values)
            current_row += 1

        output_path = normalize_xlsx_output_path(output_path)
        if os.path.exists(output_path):
            os.remove(output_path)
        dst_wb.SaveAs(os.path.abspath(output_path), FileFormat=XL_OPENXML_WORKBOOK)
        return len(selected_sheet_names)
    finally:
        if src_wb is not None:
            src_wb.Close(SaveChanges=False)
        if dst_wb is not None:
            dst_wb.Close(SaveChanges=False)
        if excel is not None:
            try:
                excel.CutCopyMode = False
            except Exception:
                pass
            excel.Quit()


def merge_multiple_files_excel_com(
    file_list,
    output_path,
    keep_all_headers=True,
    paste_values=False,
    header_row_count=1,
):
    if not file_list:
        return 0

    excel = None
    dst_wb = None
    opened_wbs = []
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        dst_wb, dst_ws = prepare_destination_workbook(excel, "Merged_Files")

        current_row = 1
        source_index = 0
        for file_path in file_list:
            converted_path = ensure_xlsx(file_path)
            src_wb = excel.Workbooks.Open(os.path.abspath(converted_path), ReadOnly=True)
            opened_wbs.append(src_wb)
            for sheet_idx in range(1, src_wb.Worksheets.Count + 1):
                source_ws = src_wb.Worksheets(sheet_idx)
                include_header = keep_all_headers or source_index == 0
                source_start_row = 1 if include_header else header_row_count + 1
                current_row = copy_excel_range(source_ws, dst_ws, source_start_row, current_row, paste_values)
                current_row += 1
                source_index += 1

        output_path = normalize_xlsx_output_path(output_path)
        if os.path.exists(output_path):
            os.remove(output_path)
        dst_wb.SaveAs(os.path.abspath(output_path), FileFormat=XL_OPENXML_WORKBOOK)
        return len(file_list)
    finally:
        for wb in opened_wbs:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if dst_wb is not None:
            dst_wb.Close(SaveChanges=False)
        if excel is not None:
            try:
                excel.CutCopyMode = False
            except Exception:
                pass
            excel.Quit()


def merge_selected_sheets_openpyxl(input_path, output_path, selected_sheet_names, keep_all_headers=True, header_row_count=1):
    if not selected_sheet_names:
        return 0

    converted_path = ensure_xlsx(input_path)
    src_wb = load_workbook(converted_path, data_only=False)
    dst_wb = Workbook()
    del dst_wb["Sheet"]
    dst_ws = dst_wb.create_sheet(title="Merged_Sheets")

    current_row = 1
    for index, sheet_name in enumerate(selected_sheet_names):
        source_ws = src_wb[sheet_name]
        include_header = keep_all_headers or index == 0
        current_row = copy_sheet_data(source_ws, dst_ws, current_row, include_header, header_row_count)
        current_row += 1

    dst_wb.save(normalize_xlsx_output_path(output_path))
    return len(selected_sheet_names)


def merge_multiple_files_openpyxl(file_list, output_path, keep_all_headers=True, header_row_count=1):
    if not file_list:
        return 0

    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = "Merged_Files"
    current_row = 1
    source_index = 0

    for file_path in file_list:
        converted_path = ensure_xlsx(file_path)
        src_wb = load_workbook(converted_path, data_only=False)
        for sheet_name in src_wb.sheetnames:
            source_ws = src_wb[sheet_name]
            include_header = keep_all_headers or source_index == 0
            current_row = copy_sheet_data(source_ws, dst_ws, current_row, include_header, header_row_count)
            current_row += 1
            source_index += 1

    dst_wb.save(normalize_xlsx_output_path(output_path))
    return len(file_list)


def merge_selected_sheets(
    input_path,
    output_path,
    selected_sheet_names,
    keep_all_headers=True,
    paste_values=False,
    prefer_excel_com=True,
    header_row_count=1,
):
    if prefer_excel_com:
        try:
            return merge_selected_sheets_excel_com(
                input_path, output_path, selected_sheet_names, keep_all_headers, paste_values, header_row_count
            )
        except Exception as exc:
            if paste_values:
                raise
            print(f"Excel COM failed, falling back to openpyxl: {exc}")
    return merge_selected_sheets_openpyxl(input_path, output_path, selected_sheet_names, keep_all_headers, header_row_count)


def merge_multiple_files(
    file_list,
    output_path,
    keep_all_headers=True,
    paste_values=False,
    prefer_excel_com=True,
    header_row_count=1,
):
    if prefer_excel_com:
        try:
            return merge_multiple_files_excel_com(
                file_list, output_path, keep_all_headers, paste_values, header_row_count
            )
        except Exception as exc:
            if paste_values:
                raise
            print(f"Excel COM failed, falling back to openpyxl: {exc}")
    return merge_multiple_files_openpyxl(file_list, output_path, keep_all_headers, header_row_count)


def copy_header_and_format(source_ws, target_ws, header_start, header_end):
    for col_letter, col_dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dimension.width
        target_ws.column_dimensions[col_letter].hidden = col_dimension.hidden

    target_row_idx = 1
    for row_idx in range(header_start, header_end + 1):
        if row_idx in source_ws.row_dimensions:
            target_ws.row_dimensions[target_row_idx].height = source_ws.row_dimensions[row_idx].height
            target_ws.row_dimensions[target_row_idx].hidden = source_ws.row_dimensions[row_idx].hidden

        for cell in source_ws[row_idx]:
            new_cell = target_ws.cell(row=target_row_idx, column=cell.column, value=cell.value)
            copy_cell_style(cell, new_cell)
        target_row_idx += 1

    for merged_range in source_ws.merged_cells.ranges:
        if merged_range.min_row >= header_start and merged_range.max_row <= header_end:
            new_min_row = merged_range.min_row - header_start + 1
            new_max_row = merged_range.max_row - header_start + 1
            new_range = (
                f"{openpyxl_get_col_letter(merged_range.min_col)}{new_min_row}:"
                f"{openpyxl_get_col_letter(merged_range.max_col)}{new_max_row}"
            )
            try:
                target_ws.merge_cells(new_range)
            except Exception:
                pass


def copy_data_chunk(source_ws, target_ws, rows_chunk, header_count):
    current_target_row = header_count + 1
    row_mapping = {}

    for row in rows_chunk:
        source_row_idx = row[0].row
        row_mapping[source_row_idx] = current_target_row

        if source_row_idx in source_ws.row_dimensions:
            target_ws.row_dimensions[current_target_row].height = source_ws.row_dimensions[source_row_idx].height
            target_ws.row_dimensions[current_target_row].hidden = source_ws.row_dimensions[source_row_idx].hidden

        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            new_cell = target_ws.cell(row=current_target_row, column=cell.column, value=cell.value)
            copy_cell_style(cell, new_cell)
        current_target_row += 1

    for merged_range in source_ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        if min_row in row_mapping and max_row in row_mapping:
            new_range = (
                f"{openpyxl_get_col_letter(merged_range.min_col)}{row_mapping[min_row]}:"
                f"{openpyxl_get_col_letter(merged_range.max_col)}{row_mapping[max_row]}"
            )
            try:
                target_ws.merge_cells(new_range)
            except Exception:
                pass


def get_last_nonempty_row_openpyxl(ws, start_row=1):
    for row_idx in range(ws.max_row, start_row - 1, -1):
        for cell in ws[row_idx]:
            if cell.value is not None and str(cell.value).strip() != "":
                return row_idx
    return start_row - 1


def copy_row_range_openpyxl(source_ws, target_ws, source_start_row, source_end_row, target_start_row):
    if source_end_row < source_start_row:
        return target_start_row - 1

    row_mapping = {}
    current_target_row = target_start_row
    for source_row_idx in range(source_start_row, source_end_row + 1):
        row_mapping[source_row_idx] = current_target_row

        if source_row_idx in source_ws.row_dimensions:
            target_ws.row_dimensions[current_target_row].height = source_ws.row_dimensions[source_row_idx].height
            target_ws.row_dimensions[current_target_row].hidden = source_ws.row_dimensions[source_row_idx].hidden

        for cell in source_ws[source_row_idx]:
            if cell.value is None and not cell.has_style:
                continue
            new_cell = target_ws.cell(row=current_target_row, column=cell.column, value=cell.value)
            copy_cell_style(cell, new_cell)
        current_target_row += 1

    for merged_range in source_ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        if min_row in row_mapping and max_row in row_mapping:
            new_range = (
                f"{openpyxl_get_col_letter(merged_range.min_col)}{row_mapping[min_row]}:"
                f"{openpyxl_get_col_letter(merged_range.max_col)}{row_mapping[max_row]}"
            )
            try:
                target_ws.merge_cells(new_range)
            except Exception:
                pass

    return current_target_row - 1


def merge_selected_sheets_to_summary_openpyxl(
    input_path,
    output_path,
    selected_sheet_names,
    header_start_row,
    header_end_row,
    data_start_row,
    data_end_row=0,
    renumber_first_column=True,
    summary_sheet_name="TongHop",
):
    if not selected_sheet_names:
        return 0

    converted_path = ensure_xlsx(input_path)
    wb = load_workbook(converted_path, data_only=False)

    if summary_sheet_name in wb.sheetnames:
        del wb[summary_sheet_name]
    summary_ws = wb.create_sheet(title=summary_sheet_name, index=0)

    first_source_ws = wb[selected_sheet_names[0]]
    for col_letter, col_dimension in first_source_ws.column_dimensions.items():
        summary_ws.column_dimensions[col_letter].width = col_dimension.width
        summary_ws.column_dimensions[col_letter].hidden = col_dimension.hidden

    current_row = copy_row_range_openpyxl(first_source_ws, summary_ws, header_start_row, header_end_row, 1) + 1

    sequence_number = 1
    for sheet_name in selected_sheet_names:
        source_ws = wb[sheet_name]
        source_end_row = data_end_row if data_end_row and data_end_row >= data_start_row else get_last_nonempty_row_openpyxl(
            source_ws, data_start_row
        )
        if source_end_row < data_start_row:
            continue

        target_start_row = current_row
        current_row = copy_row_range_openpyxl(source_ws, summary_ws, data_start_row, source_end_row, target_start_row)
        if renumber_first_column:
            for row_idx in range(target_start_row, current_row + 1):
                summary_ws.cell(row=row_idx, column=1).value = sequence_number
                sequence_number += 1
        current_row += 1

    wb.save(normalize_xlsx_output_path(output_path))
    wb.close()
    return len(selected_sheet_names)


def merge_selected_sheets_to_summary(
    input_path,
    output_path,
    selected_sheet_names,
    header_start_row,
    header_end_row,
    data_start_row,
    data_end_row=0,
    paste_values=False,
    prefer_excel_com=True,
    renumber_first_column=True,
    summary_sheet_name="TongHop",
):
    if prefer_excel_com:
        try:
            return merge_selected_sheets_to_summary_excel_com(
                input_path,
                output_path,
                selected_sheet_names,
                header_start_row,
                header_end_row,
                data_start_row,
                data_end_row,
                paste_values,
                renumber_first_column,
                summary_sheet_name,
            )
        except Exception as exc:
            if paste_values:
                raise
            print(f"Excel COM failed, falling back to openpyxl: {exc}")

    return merge_selected_sheets_to_summary_openpyxl(
        input_path,
        output_path,
        selected_sheet_names,
        header_start_row,
        header_end_row,
        data_start_row,
        data_end_row,
        renumber_first_column,
        summary_sheet_name,
    )


def clean_sheet_name(name):
    sheet_name = str(name)[:30].strip()
    for char in ["/", "\\", "?", "*", ":", "[", "]"]:
        sheet_name = sheet_name.replace(char, "")
    return sheet_name or "Sheet"


def split_excel(
    input_path,
    output_path_base,
    mode,
    output_mode,
    header_start,
    header_end,
    sheet_name=None,
    column_index=None,
    row_count=None,
):
    converted_path = ensure_xlsx(input_path)
    wb = load_workbook(converted_path, data_only=False)
    if not wb.sheetnames:
        return 0

    source_ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    real_max_row = get_real_max_row(source_ws)
    header_count = header_end - header_start + 1
    chunks = {}

    if mode == "column":
        for row in source_ws.iter_rows(min_row=header_end + 1, max_row=real_max_row):
            value = row[column_index].value
            key = str(value) if value is not None else "Blank"
            chunks.setdefault(key, []).append(row)
    else:
        current_chunk = 1
        current_rows = []
        for row in source_ws.iter_rows(min_row=header_end + 1, max_row=real_max_row):
            current_rows.append(row)
            if len(current_rows) >= row_count:
                chunks[f"Part_{current_chunk}"] = current_rows
                current_chunk += 1
                current_rows = []
        if current_rows:
            chunks[f"Part_{current_chunk}"] = current_rows

    if not chunks:
        return 0

    if output_mode == "sheets":
        for name, rows in chunks.items():
            safe_name = clean_sheet_name(name)
            base_name = safe_name
            counter = 1
            while safe_name in wb.sheetnames:
                safe_name = f"{base_name}_{counter}"
                counter += 1
            new_ws = wb.create_sheet(title=safe_name)
            copy_header_and_format(source_ws, new_ws, header_start, header_end)
            copy_data_chunk(source_ws, new_ws, rows, header_count)
        wb.save(normalize_xlsx_output_path(output_path_base))
        return len(chunks)

    base_dir = os.path.dirname(output_path_base)
    base_name_file, ext = os.path.splitext(os.path.basename(output_path_base))
    ext = ext or ".xlsx"
    for name, rows in chunks.items():
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = clean_sheet_name(name)
        copy_header_and_format(source_ws, new_ws, header_start, header_end)
        copy_data_chunk(source_ws, new_ws, rows, header_count)
        out_file = os.path.join(base_dir, f"{base_name_file}_{clean_sheet_name(name)}{ext}")
        new_wb.save(out_file)
    return len(chunks)


class ExcelMergerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Excel Tool - Ghep Sheet/File")
        self.geometry("1100x700")

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self.navigation_frame = ctk.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)

        self.label_header = ctk.CTkLabel(
            self.navigation_frame,
            text="  EXCEL TOOLBOX",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=("black", "white"),
        )
        self.label_header.grid(row=0, column=0, padx=20, pady=20)

        self.home_button = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="Ghep Sheet (1 File)",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            command=lambda: self.select_frame_by_name("merge_sheet"),
        )
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.frame_2_button = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="Ghep Nhieu File",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            command=lambda: self.select_frame_by_name("merge_files"),
        )
        self.frame_2_button.grid(row=2, column=0, sticky="ew")

        self.frame_3_button = ctk.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="Tach File Excel",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            anchor="w",
            command=lambda: self.select_frame_by_name("split_file"),
        )
        self.frame_3_button.grid(row=3, column=0, sticky="ew")

        self.home_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.second_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.third_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")

        self.setup_home_frame()
        self.setup_second_frame()
        self.setup_third_frame()
        self.select_frame_by_name("merge_sheet")

    def setup_merge_options(self, parent, row, column=0, columnspan=1):
        options = ctk.CTkFrame(parent, fg_color="transparent")
        options.grid(row=row, column=column, columnspan=columnspan, padx=20, pady=5, sticky="w")
        keep_headers_var = ctk.BooleanVar(value=True)
        paste_values_var = ctk.BooleanVar(value=False)
        prefer_com_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(options, text="Giu header cua moi sheet/file", variable=keep_headers_var).grid(row=0, column=0, padx=8)
        ctk.CTkCheckBox(options, text="Paste gia tri thay vi cong thuc", variable=paste_values_var).grid(row=0, column=1, padx=8)
        ctk.CTkCheckBox(options, text="Uu tien Excel COM de giu format tot nhat", variable=prefer_com_var).grid(row=0, column=2, padx=8)
        return keep_headers_var, paste_values_var, prefer_com_var

    def setup_summary_merge_options(self, parent, row, column=0, columnspan=1):
        options = ctk.CTkFrame(parent, fg_color="transparent")
        options.grid(row=row, column=column, columnspan=columnspan, padx=20, pady=5, sticky="w")
        paste_values_var = ctk.BooleanVar(value=False)
        prefer_com_var = ctk.BooleanVar(value=True)
        renumber_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(options, text="Paste gia tri thay vi cong thuc", variable=paste_values_var).grid(row=0, column=0, padx=8)
        ctk.CTkCheckBox(options, text="Uu tien Excel COM de giu format tot nhat", variable=prefer_com_var).grid(row=0, column=1, padx=8)
        ctk.CTkCheckBox(options, text="Danh lai STT cot A", variable=renumber_var).grid(row=0, column=2, padx=8)
        return paste_values_var, prefer_com_var, renumber_var

    def setup_home_frame(self):
        ctk.CTkLabel(
            self.home_frame,
            text="Ghep Sheet (1 File) vao TongHop",
            font=ctk.CTkFont(size=20, weight="bold"),
        ).grid(row=0, column=0, padx=20, pady=20)
        self.input_file_path = ctk.StringVar()

        file_frame = ctk.CTkFrame(self.home_frame, fg_color="transparent")
        file_frame.grid(row=1, column=0, padx=20, pady=10)
        ctk.CTkEntry(file_frame, textvariable=self.input_file_path, width=400).grid(row=0, column=0, padx=5)
        ctk.CTkButton(file_frame, text="Chon File", command=self.browse_single_file).grid(row=0, column=1, padx=5)

        self.sheet_info_label = ctk.CTkLabel(self.home_frame, text="Chon cac sheet can ghep:", text_color=("gray40", "gray60"))
        self.sheet_info_label.grid(row=2, column=0, pady=(10, 0))

        list_frame = ctk.CTkFrame(self.home_frame)
        list_frame.grid(row=3, column=0, padx=20, pady=5)
        self.home_sheet_listbox = tk.Listbox(list_frame, selectmode="multiple", width=60, height=8, bg="#333333", fg="white", selectbackground="#2ecc71")
        self.home_sheet_listbox.pack(side="left", fill="both", expand=True)
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=self.home_sheet_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.home_sheet_listbox.config(yscrollcommand=scrollbar.set)

        ctk.CTkButton(self.home_frame, text="Chon tat ca Sheet", command=self.select_all_home_sheets, width=150).grid(row=4, column=0, pady=5)

        range_frame = ctk.CTkFrame(self.home_frame, fg_color="transparent")
        range_frame.grid(row=5, column=0, padx=20, pady=5)
        ctk.CTkLabel(range_frame, text="Tieu de tu dong:").grid(row=0, column=0, padx=2)
        self.merge_header_start_entry = ctk.CTkEntry(range_frame, width=50)
        self.merge_header_start_entry.insert(0, "1")
        self.merge_header_start_entry.grid(row=0, column=1, padx=2)
        ctk.CTkLabel(range_frame, text="den:").grid(row=0, column=2, padx=2)
        self.merge_header_end_entry = ctk.CTkEntry(range_frame, width=50)
        self.merge_header_end_entry.insert(0, "7")
        self.merge_header_end_entry.grid(row=0, column=3, padx=2)
        ctk.CTkLabel(range_frame, text="Du lieu tu dong:").grid(row=0, column=4, padx=(16, 2))
        self.merge_data_start_entry = ctk.CTkEntry(range_frame, width=50)
        self.merge_data_start_entry.insert(0, "8")
        self.merge_data_start_entry.grid(row=0, column=5, padx=2)
        ctk.CTkLabel(range_frame, text="den:").grid(row=0, column=6, padx=2)
        self.merge_data_end_entry = ctk.CTkEntry(range_frame, width=70)
        self.merge_data_end_entry.insert(0, "0")
        self.merge_data_end_entry.grid(row=0, column=7, padx=2)
        ctk.CTkLabel(range_frame, text="0 = tu dong dong cuoi co du lieu").grid(row=0, column=8, padx=8)

        self.home_paste_values_var, self.home_prefer_com_var, self.home_renumber_var = self.setup_summary_merge_options(self.home_frame, 6)
        ctk.CTkButton(self.home_frame, text="Bat dau Ghep Sheet", command=self.merge_sheets_process, height=50, fg_color="#2ecc71").grid(row=7, column=0, pady=20)

    def setup_second_frame(self):
        ctk.CTkLabel(
            self.second_frame,
            text="Ghep nhieu file Excel - uu tien giu nguyen dinh dang bang Excel",
            font=ctk.CTkFont(size=20, weight="bold"),
        ).grid(row=0, column=0, padx=20, pady=20)
        self.file_list = []
        self.file_listbox = tk.Listbox(self.second_frame, width=80, height=10, bg="#333333", fg="white")
        self.file_listbox.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        file_ops = ctk.CTkFrame(self.second_frame, fg_color="transparent")
        file_ops.grid(row=1, column=1, padx=20, pady=10, sticky="n")
        ctk.CTkButton(file_ops, text="Them Files", command=self.browse_multiple_files).grid(row=0, column=0, pady=5)
        ctk.CTkButton(file_ops, text="Xoa", fg_color="#e74c3c", command=self.clear_file_list).grid(row=1, column=0, pady=5)
        self.files_keep_headers_var, self.files_paste_values_var, self.files_prefer_com_var = self.setup_merge_options(self.second_frame, 2, column=0, columnspan=2)
        ctk.CTkButton(self.second_frame, text="Bat dau Ghep File", command=self.merge_files_process, height=50, fg_color="#3498db").grid(row=3, column=0, columnspan=2, pady=30)

    def setup_third_frame(self):
        ctk.CTkLabel(self.third_frame, text="Tach File Excel", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, columnspan=2, padx=20, pady=20)
        self.split_input_path = ctk.StringVar()
        file_f = ctk.CTkFrame(self.third_frame, fg_color="transparent")
        file_f.grid(row=1, column=0, columnspan=2, padx=20, pady=10)
        ctk.CTkEntry(file_f, textvariable=self.split_input_path, width=400).grid(row=0, column=0, padx=5)
        ctk.CTkButton(file_f, text="Chon File", command=self.browse_split_file).grid(row=0, column=1, padx=5)

        sheet_f = ctk.CTkFrame(self.third_frame, fg_color="transparent")
        sheet_f.grid(row=2, column=0, columnspan=2, padx=20, pady=5)
        ctk.CTkLabel(sheet_f, text="Chon Sheet can tach:").grid(row=0, column=0, padx=5)
        self.split_sheet_var = ctk.StringVar(value="Active Sheet")
        self.split_sheet_menu = ctk.CTkOptionMenu(sheet_f, variable=self.split_sheet_var, values=["Active Sheet"], command=lambda _: self.reload_columns())
        self.split_sheet_menu.grid(row=0, column=1, padx=5)

        options_frame = ctk.CTkFrame(self.third_frame, fg_color="transparent")
        options_frame.grid(row=3, column=0, columnspan=2, pady=10)

        self.split_mode_var = ctk.StringVar(value="column")
        mode_frame = ctk.CTkFrame(options_frame)
        mode_frame.grid(row=0, column=0, padx=10, sticky="n")
        ctk.CTkLabel(mode_frame, text="Phuong thuc tach:", font=ctk.CTkFont(weight="bold")).pack(pady=5)
        ctk.CTkRadioButton(mode_frame, text="Theo gia tri cot", variable=self.split_mode_var, value="column", command=self.toggle_split_inputs).pack(anchor="w", padx=10, pady=5)
        ctk.CTkRadioButton(mode_frame, text="Theo so luong dong", variable=self.split_mode_var, value="row_count", command=self.toggle_split_inputs).pack(anchor="w", padx=10, pady=5)

        self.split_output_var = ctk.StringVar(value="sheets")
        output_frame = ctk.CTkFrame(options_frame)
        output_frame.grid(row=0, column=1, padx=10, sticky="n")
        ctk.CTkLabel(output_frame, text="Hinh thuc xuat:", font=ctk.CTkFont(weight="bold")).pack(pady=5)
        ctk.CTkRadioButton(output_frame, text="Nhieu Sheet trong 1 File", variable=self.split_output_var, value="sheets").pack(anchor="w", padx=10, pady=5)
        ctk.CTkRadioButton(output_frame, text="Nhieu File rieng biet", variable=self.split_output_var, value="files").pack(anchor="w", padx=10, pady=5)

        self.inputs_frame = ctk.CTkFrame(self.third_frame, fg_color="transparent")
        self.inputs_frame.grid(row=4, column=0, columnspan=2, pady=5)
        ctk.CTkLabel(self.inputs_frame, text="Tieu de tu dong:").grid(row=0, column=0, padx=2)
        self.header_start_entry = ctk.CTkEntry(self.inputs_frame, width=40)
        self.header_start_entry.insert(0, "1")
        self.header_start_entry.grid(row=0, column=1, padx=2)
        ctk.CTkLabel(self.inputs_frame, text="den:").grid(row=0, column=2, padx=2)
        self.header_end_entry = ctk.CTkEntry(self.inputs_frame, width=40)
        self.header_end_entry.insert(0, "1")
        self.header_end_entry.grid(row=0, column=3, padx=2)

        self.split_header_menu = ctk.CTkOptionMenu(self.inputs_frame, values=["Chua co du lieu"], width=180)
        self.split_header_menu.grid(row=0, column=4, padx=5)
        self.row_count_entry = ctk.CTkEntry(self.inputs_frame, width=150, placeholder_text="So dong moi phan...")
        self.row_count_entry.grid_remove()
        ctk.CTkButton(self.inputs_frame, text="Tai lai cot", width=80, command=self.reload_columns).grid(row=0, column=5, padx=5)
        ctk.CTkButton(self.third_frame, text="Bat dau Tach File", command=self.split_process, height=50, fg_color="#9b59b6").grid(row=5, column=0, columnspan=2, pady=30)

    def select_all_home_sheets(self):
        self.home_sheet_listbox.select_set(0, tk.END)

    def toggle_split_inputs(self):
        if self.split_mode_var.get() == "column":
            self.split_header_menu.grid()
            self.row_count_entry.grid_remove()
        else:
            self.split_header_menu.grid_remove()
            self.row_count_entry.grid(row=0, column=4, padx=5)

    def select_frame_by_name(self, name):
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "merge_sheet" else "transparent")
        self.frame_2_button.configure(fg_color=("gray75", "gray25") if name == "merge_files" else "transparent")
        self.frame_3_button.configure(fg_color=("gray75", "gray25") if name == "split_file" else "transparent")

        self.home_frame.grid_forget()
        self.second_frame.grid_forget()
        self.third_frame.grid_forget()
        if name == "merge_sheet":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        elif name == "merge_files":
            self.second_frame.grid(row=0, column=1, sticky="nsew")
        elif name == "split_file":
            self.third_frame.grid(row=0, column=1, sticky="nsew")

    def browse_single_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        self.input_file_path.set(file_path)
        try:
            sheet_names = get_workbook_sheet_names(file_path)
            self.home_sheet_listbox.delete(0, tk.END)
            for name in sheet_names:
                self.home_sheet_listbox.insert(tk.END, name)
            self.home_sheet_listbox.select_set(0, tk.END)
        except Exception as exc:
            messagebox.showerror("Loi", str(exc))

    def browse_split_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        self.split_input_path.set(file_path)
        try:
            sheet_names = get_workbook_sheet_names(file_path)
            self.split_sheet_menu.configure(values=sheet_names)
            if sheet_names:
                self.split_sheet_var.set(sheet_names[0])
            self.reload_columns()
        except Exception as exc:
            messagebox.showerror("Loi", str(exc))

    def reload_columns(self):
        file_path = self.split_input_path.get()
        if not file_path or not os.path.exists(file_path):
            return
        try:
            header_end = int(self.header_end_entry.get())
            selected_sheet = self.split_sheet_var.get()
            sheet_name = selected_sheet if selected_sheet != "Active Sheet" else 0
            values = get_header_columns(file_path, sheet_name, header_end)
            self.split_header_menu.configure(values=values or ["Chua co du lieu"])
            if values:
                self.split_header_menu.set(values[0])
        except Exception as exc:
            messagebox.showerror("Loi", str(exc))

    def split_process(self):
        input_file = self.split_input_path.get()
        if not input_file:
            messagebox.showwarning("Canh bao", "Vui long chon file can tach!")
            return

        try:
            header_start = int(self.header_start_entry.get())
            header_end = int(self.header_end_entry.get())
            if header_start < 1 or header_end < header_start:
                messagebox.showwarning("Canh bao", "Dai dong tieu de khong hop le!")
                return

            mode = self.split_mode_var.get()
            output_mode = self.split_output_var.get()
            selected_sheet = self.split_sheet_var.get()
            row_count = None
            column_index = None

            if mode == "column":
                selected_value = self.split_header_menu.get()
                if selected_value == "Chua co du lieu" or not selected_value:
                    messagebox.showwarning("Canh bao", "Vui long tai lai va chon cot de tach!")
                    return
                column_index = get_column_index(selected_value.split(" | ")[0])
            else:
                row_count = int(self.row_count_entry.get())
                if row_count <= 0:
                    raise ValueError

            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if not save_path:
                return

            num_parts = split_excel(
                input_file,
                save_path,
                mode,
                output_mode,
                header_start,
                header_end,
                sheet_name=selected_sheet,
                column_index=column_index,
                row_count=row_count,
            )
            messagebox.showinfo("Thanh cong", f"Da tach xong thanh {num_parts} phan!")
        except Exception as exc:
            messagebox.showerror("Loi", f"Co loi xay ra: {exc}")

    def merge_sheets_process(self):
        input_path = self.input_file_path.get()
        if not input_path:
            messagebox.showwarning("Canh bao", "Vui long chon file!")
            return
        selected_indices = self.home_sheet_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Canh bao", "Vui long chon it nhat 1 sheet de ghep!")
            return

        selected_sheets = [self.home_sheet_listbox.get(i) for i in selected_indices]
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:
            return

        try:
            header_start = int(self.merge_header_start_entry.get())
            header_end = int(self.merge_header_end_entry.get())
            data_start = int(self.merge_data_start_entry.get())
            data_end = int(self.merge_data_end_entry.get())
            if header_start < 1 or header_end < header_start:
                messagebox.showwarning("Canh bao", "Dai dong tieu de khong hop le!")
                return
            if data_start < 1 or (data_end != 0 and data_end < data_start):
                messagebox.showwarning("Canh bao", "Dai dong du lieu khong hop le! Nhap 0 o dong ket thuc de tu dong.")
                return

            count = merge_selected_sheets_to_summary(
                input_path,
                output_path,
                selected_sheets,
                header_start,
                header_end,
                data_start,
                data_end,
                paste_values=self.home_paste_values_var.get(),
                prefer_excel_com=self.home_prefer_com_var.get(),
                renumber_first_column=self.home_renumber_var.get(),
                summary_sheet_name="TongHop",
            )
            messagebox.showinfo("Thanh cong", f"Da ghep xong {count} sheet vao TongHop!")
        except Exception as exc:
            messagebox.showerror("Loi", str(exc))

    def browse_multiple_files(self):
        for file_path in filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")]):
            if file_path not in self.file_list:
                self.file_list.append(file_path)
                self.file_listbox.insert(tk.END, os.path.basename(file_path))

    def clear_file_list(self):
        self.file_list = []
        self.file_listbox.delete(0, tk.END)

    def merge_files_process(self):
        if not self.file_list:
            messagebox.showwarning("Canh bao", "Vui long them file can ghep!")
            return
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:
            return
        try:
            count = merge_multiple_files(
                self.file_list,
                output_path,
                keep_all_headers=self.files_keep_headers_var.get(),
                paste_values=self.files_paste_values_var.get(),
                prefer_excel_com=self.files_prefer_com_var.get(),
            )
            messagebox.showinfo("Thanh cong", f"Da ghep xong {count} file!")
        except Exception as exc:
            messagebox.showerror("Loi", str(exc))


if __name__ == "__main__":
    app = ExcelMergerApp()
    app.mainloop()
